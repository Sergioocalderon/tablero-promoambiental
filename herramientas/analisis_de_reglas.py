"""
validar_reglas_geotab.py
=========================
Valida reglas de Geotab (creadas por el equipo) contra datos reales, respondiendo:

  1. ¿Cuántos eventos de excepción disparó cada regla en el periodo analizado?
     (soporte real de datos: una regla que nunca dispara puede tener un error de lógica,
     como la de IsValueLessThan[value=0](IsDriving) que revisamos).

  2. ¿Cuántas muestras de StatusData hay disponibles para los diagnósticos que usa
     cada regla (RPM, Speed, PowerTakeoffEngaged, etc.), y con qué frecuencia
     (gap mediano en segundos)? Esto permite saber si un DurationLongerThan corto
     (ej. 3s) es alcanzable con la cadencia real de telemetría del vehículo, o si
     nunca se cumple porque el gap real es de varios minutos.

  3. ¿En cuántos vehículos distintos disparó cada regla? Ayuda a detectar reglas que
     solo afectan a 1-2 unidades (posible caso puntual, no patrón de flota) vs. reglas
     que disparan en toda la flota (posible umbral mal calibrado / falso positivo masivo).

Requisitos:
    pip install mygeotab pandas python-dotenv

Credenciales:
    El script busca primero un archivo .env en la misma carpeta con estas variables
    (mismos nombres que ya usas en tus otros scripts del proyecto):
        GEOTAB_USUARIO=...
        GEOTAB_CONTRASENA=...
        GEOTAB_DATABASE=...
        GEOTAB_SERVER=...
    Si falta alguna, el script la pide por consola.

Uso básico (valida reglas específicas por nombre, o parte del nombre):
    python validar_reglas_geotab.py --start 2026-07-01 --end 2026-08-14 \
        --rules "RALENTI TOTAL" "REVOLUCIONES ALTAS CON PTO"

Uso sin filtro de nombre (valida TODAS las reglas de la base de datos):
    python validar_reglas_geotab.py --start 2026-07-01 --end 2026-08-14

Para además revisar el gap de muestreo de telemetría en vehículos puntuales:
    python validar_reglas_geotab.py --start 2026-07-01 --end 2026-08-14 \
        --rules "RALENTI TOTAL" --devices b1234 b5678

(Los device IDs se obtienen de la columna 'vehiculos_involucrados' del CSV de salida,
 o desde MyGeotab > Vehículos > exportar lista.)

Salida: imprime un resumen en consola y guarda un CSV consolidado
(por defecto: validacion_reglas.csv) con una fila por regla.
"""

import argparse
import getpass
import os
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta, timezone

import mygeotab
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from dotenv import load_dotenv
    load_dotenv()  # carga el .env de la carpeta actual, si existe
except ImportError:
    pass  # si no está instalado python-dotenv, simplemente se piden las credenciales a mano


# ---------------------------------------------------------------------------
# Utilidad de reintento (para 504 Gateway Timeout y errores de red transitorios)
# ---------------------------------------------------------------------------

def _get_con_reintento(api, type_name, search, max_intentos=4, espera_base=5):
    ultimo_error = None
    for intento in range(1, max_intentos + 1):
        try:
            return api.get(type_name, search=search)
        except Exception as e:
            ultimo_error = e
            if intento == max_intentos:
                break
            espera = espera_base * intento
            print(f"      (intento {intento}/{max_intentos} falló: {e}. Reintentando en {espera}s...)")
            time.sleep(espera)
    print(f"      *** No se pudo completar la consulta tras {max_intentos} intentos: {ultimo_error} ***")
    return None


# ---------------------------------------------------------------------------
# Conexión
# ---------------------------------------------------------------------------

def conectar():
    server = os.getenv("GEOTAB_SERVER") or input("Servidor Geotab (ej. my.geotab.com) [Enter = my.geotab.com]: ").strip() or "my.geotab.com"
    database = os.getenv("GEOTAB_DATABASE") or input("Base de datos (Database name en MyGeotab): ").strip()
    username = os.getenv("GEOTAB_USUARIO") or input("Usuario (correo): ").strip()
    password = os.getenv("GEOTAB_CONTRASENA") or getpass.getpass("Contraseña: ")

    if not password:
        password = getpass.getpass(f"GEOTAB_CONTRASENA está vacía en .env. Escribe la contraseña para '{username}': ")

    print(f"Conectando como '{username}' a la base de datos '{database}' en {server}...")

    api = mygeotab.API(username=username, password=password, database=database, server=server)
    try:
        api.authenticate()
    except mygeotab.exceptions.AuthenticationException as e:
        print(f"\nError de autenticación: {e}")
        sys.exit(1)

    print(f"Conectado correctamente.\n")
    cargar_mapa_placas(api)
    return api


# ---------------------------------------------------------------------------
# Reglas
# ---------------------------------------------------------------------------

def obtener_reglas(api, nombres_buscados, solo_custom=False):
    """Trae todas las reglas y, si se dan nombres, filtra por coincidencia parcial (sin distinguir mayúsculas).
    Si solo_custom=True, excluye las reglas con baseType == 'Stock' (predefinidas de Geotab)."""
    todas = api.get("Rule")

    if solo_custom:
        todas = [r for r in todas if (r.get("baseType") or "").lower() != "stock"]

    if not nombres_buscados:
        return todas

    nombres_lower = [_normalizar(n) for n in nombres_buscados]
    filtradas = [
        r for r in todas
        if any(nb in _normalizar(r.get("name") or "") for nb in nombres_lower)
    ]
    return filtradas


def _normalizar(texto):
    """Minúsculas y sin tildes, para que 'RALENTI' encuentre 'RALENTÍ'."""
    import unicodedata
    texto = texto.lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c))


def extraer_diagnosticos_de_condicion(condition, encontrados=None):
    """Recorre recursivamente el árbol de condiciones (And/Or/anidados) y extrae los diagnostic id usados."""
    if encontrados is None:
        encontrados = set()
    if not condition:
        return encontrados

    diag = condition.get("diagnostic")
    if diag and diag.get("id"):
        encontrados.add(diag["id"])

    for hijo in (condition.get("children") or []):
        extraer_diagnosticos_de_condicion(hijo, encontrados)

    return encontrados


# ---------------------------------------------------------------------------
# Eventos de excepción (soporte real de datos de cada regla)
# ---------------------------------------------------------------------------

def contar_eventos_excepcion(api, rule, fecha_inicio, fecha_fin, dias_por_bloque=7):
    """Consulta ExceptionEvent en bloques (por defecto semanales) para evitar timeouts del servidor
    en reglas con mucho volumen (ej. ralentí), y reintenta automáticamente ante fallos transitorios."""
    total = []
    cursor = fecha_inicio
    while cursor < fecha_fin:
        siguiente = min(cursor + timedelta(days=dias_por_bloque), fecha_fin)
        search = {
            "ruleSearch": {"id": rule["id"]},
            "fromDate": cursor.isoformat(),
            "toDate": siguiente.isoformat(),
        }
        eventos = _get_con_reintento(api, "ExceptionEvent", search)
        if eventos:
            total.extend(eventos)
        cursor = siguiente
    return total


def resumen_por_dispositivo(eventos):
    conteo = defaultdict(int)
    for e in eventos:
        dev = (e.get("device") or {}).get("id", "desconocido")
        conteo[dev] += 1
    return conteo


def _parsear_duracion_a_minutos(duracion_valor):
    """Convierte el campo 'duration' de un ExceptionEvent a minutos (float).
    mygeotab puede devolverlo como distintos tipos según la versión del SDK:
    - str en formato TimeSpan: '0.00:07:23.4500000' o '00:07:23.4500000'
    - datetime.time (para duraciones menores a 24h)
    - datetime.timedelta
    """
    if duracion_valor is None:
        return None

    import datetime as _dt

    if isinstance(duracion_valor, _dt.timedelta):
        return duracion_valor.total_seconds() / 60

    if isinstance(duracion_valor, _dt.time):
        return (duracion_valor.hour * 60 + duracion_valor.minute
                + duracion_valor.second / 60 + duracion_valor.microsecond / 60_000_000)

    if isinstance(duracion_valor, (int, float)):
        # Asumimos que ya viene en segundos si es numérico
        return duracion_valor / 60

    duracion_str = str(duracion_valor)
    try:
        dias = 0
        resto = duracion_str
        if "." in duracion_str.split(":")[0]:
            partes_dia = duracion_str.split(".", 1)
            dias = int(partes_dia[0])
            resto = partes_dia[1]
        partes = resto.split(":")
        horas = int(partes[0])
        minutos = int(partes[1])
        segundos = float(partes[2])
        return dias * 24 * 60 + horas * 60 + minutos + segundos / 60
    except (ValueError, IndexError):
        return None


def _normalizar_fecha_para_excel(valor, offset_horas_bogota=-5):
    """Convierte la fecha de un evento (que Geotab siempre entrega en UTC) a hora de
    Bogotá (UTC-5, Colombia no tiene horario de verano), y quita la zona horaria antes
    de escribirla en Excel. openpyxl no soporta datetimes con tzinfo: si se le pasa uno,
    wb.save() falla A MITAD de la escritura, dejando un archivo .xlsx parcial y corrupto
    en disco (en vez de no crear nada). Por eso hay que limpiar la fecha ANTES de llegar
    a esa etapa."""
    if isinstance(valor, datetime):
        if valor.tzinfo is not None:
            valor = valor.astimezone(timezone.utc).replace(tzinfo=None)
        # A partir de aquí 'valor' es naive pero representa UTC (así lo entrega Geotab
        # siempre, con o sin tzinfo explícito) -> lo convertimos a hora de Bogotá.
        valor = valor + timedelta(hours=offset_horas_bogota)
        return valor
    return valor


def analizar_duracion_eventos(eventos):
    """Dado un listado de ExceptionEvent ya descargado, calcula estadísticas de
    duración (en minutos) globales y por vehículo."""
    filas = []
    for e in eventos:
        dur_min = _parsear_duracion_a_minutos(e.get("duration"))
        if dur_min is None:
            continue
        filas.append({
            "device_id": placa_de((e.get("device") or {}).get("id", "desconocido")),
            "activeFrom": _normalizar_fecha_para_excel(e.get("activeFrom")),
            "duracion_min": dur_min,
        })

    if not filas:
        print("No se pudo extraer la duración de ningún evento (revisa el formato del campo 'duration').")
        return None

    df = pd.DataFrame(filas)

    print(f"\n=== DISTRIBUCIÓN DE DURACIÓN ({len(df)} eventos con duración válida) ===")
    print(f"  Mínimo:       {df['duracion_min'].min():.1f} min")
    print(f"  Percentil 25: {df['duracion_min'].quantile(0.25):.1f} min")
    print(f"  Mediana:      {df['duracion_min'].median():.1f} min")
    print(f"  Percentil 75: {df['duracion_min'].quantile(0.75):.1f} min")
    print(f"  Percentil 90: {df['duracion_min'].quantile(0.90):.1f} min")
    print(f"  Máximo:       {df['duracion_min'].max():.1f} min")

    # Cuántos eventos caen en distintos rangos, útil para decidir si el umbral está bien calibrado
    bins = [0, 5, 10, 15, 20, 30, 60, float("inf")]
    labels = ["<5min", "5-10min", "10-15min", "15-20min", "20-30min", "30-60min", ">60min"]
    df["rango"] = pd.cut(df["duracion_min"], bins=bins, labels=labels, right=False)
    print("\n  Distribución por rango:")
    for label, count in df["rango"].value_counts().sort_index().items():
        print(f"    {label:>10}: {count}")

    print("\n  Por vehículo (top 10 por cantidad de eventos):")
    resumen_dev = df.groupby("device_id").agg(
        eventos=("duracion_min", "count"),
        duracion_media_min=("duracion_min", "mean"),
        duracion_max_min=("duracion_min", "max"),
    ).sort_values("eventos", ascending=False)
    for dev_id, fila in resumen_dev.head(10).iterrows():
        print(f"    {dev_id}: {int(fila['eventos'])} eventos, promedio {fila['duracion_media_min']:.1f} min, máx {fila['duracion_max_min']:.1f} min")

    return df


def generar_reporte_duracion_excel(df, nombre_regla, xlsx_path):
    """Genera un Excel con el detalle de CADA evento (no solo el top 10), ordenado de mayor
    a menor duración, con semáforo de color por rango, para revisión manual uno por uno."""
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment, Border as _Border, Side as _Side
    from openpyxl.utils import get_column_letter as _get_column_letter
    from openpyxl.worksheet.table import Table as _Table, TableStyleInfo as _TableStyleInfo

    FUENTE = "Arial"
    df_ordenado = df.sort_values("duracion_min", ascending=False).reset_index(drop=True)

    wb = _Workbook()
    ws = wb.active
    ws.title = "Detalle de eventos"

    titulo_font = _Font(name=FUENTE, size=14, bold=True, color="1F4E78")
    subtitulo_font = _Font(name=FUENTE, size=10, italic=True, color="595959")
    header_font = _Font(name=FUENTE, size=11, bold=True, color="FFFFFF")
    header_fill = _PatternFill("solid", fgColor="1F4E78")

    ws["A1"] = f"Detalle de eventos — {nombre_regla}"
    ws["A1"].font = titulo_font
    ws["A2"] = f"{len(df_ordenado)} eventos, ordenados de mayor a menor duración. Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = subtitulo_font

    headers = ["Placa / Vehículo", "Inicio del evento (hora Bogotá)", "Duración (min)", "Rango"]
    fila_header = 4
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=fila_header, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = _Alignment(horizontal="center", vertical="center")

    thin = _Side(style="thin", color="D9D9D9")
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_verde = _PatternFill("solid", fgColor="C6E0B4")   # <15 min: probable espera operativa normal
    fill_amarillo = _PatternFill("solid", fgColor="FFEB9C")  # 15-30 min: revisar
    fill_rojo = _PatternFill("solid", fgColor="FFC7CE")     # >30 min: atípico, prioridad

    for i, r in df_ordenado.iterrows():
        fila = fila_header + 1 + i
        dur = r["duracion_min"]
        rango = str(r.get("rango", ""))
        valores = [r["device_id"], r["activeFrom"], round(dur, 1), rango]
        for col_idx, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col_idx, value=val)
            c.font = _Font(name=FUENTE, size=10)
            c.border = border
            c.alignment = _Alignment(horizontal="center", vertical="center")

        if dur >= 30:
            fill = fill_rojo
        elif dur >= 15:
            fill = fill_amarillo
        else:
            fill = fill_verde
        ws.cell(row=fila, column=3).fill = fill

    widths = [20, 26, 16, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[_get_column_letter(idx)].width = w

    ws.freeze_panes = f"A{fila_header + 1}"
    ultima_fila = fila_header + len(df_ordenado)
    tabla = _Table(displayName="TablaDetalleDuracion", ref=f"A{fila_header}:D{ultima_fila}")
    tabla.tableStyleInfo = _TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(tabla)

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# StatusData (volumen y gap de telemetría por diagnóstico)
# ---------------------------------------------------------------------------

def contar_statusdata(api, device_id, diagnostic_id, fecha_inicio, fecha_fin):
    search = {
        "deviceSearch": {"id": device_id},
        "diagnosticSearch": {"id": diagnostic_id},
        "fromDate": fecha_inicio.isoformat(),
        "toDate": fecha_fin.isoformat(),
    }
    return api.get("StatusData", search=search)


def gap_mediano_segundos(datos):
    if len(datos) < 2:
        return None
    fechas = sorted(pd.to_datetime([d["dateTime"] for d in datos]))
    diffs = pd.Series(fechas).diff().dropna().dt.total_seconds()
    if diffs.empty:
        return None
    return round(diffs.median(), 1)


# ---------------------------------------------------------------------------
# Reporte Excel profesional (Resumen + Detalle con semáforo de alertas)
# ---------------------------------------------------------------------------

def generar_reporte_excel(csv_path, xlsx_path, fecha_inicio_str, fecha_fin_str, fleet_size=None):
    """Convierte el CSV consolidado en un Excel con dos hojas:
    - Resumen: totales calculados con fórmulas (COUNTIF/SUM) y leyenda del semáforo.
    - Detalle: una fila por regla, con semáforo de color según:
        ROJO    -> 0 eventos en el periodo (posible error de lógica en la condición)
        AMARILLO-> disparó en >=90% de la flota (posible umbral demasiado sensible)
        VERDE   -> comportamiento normal
    """
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    df = df.sort_values("eventos_disparados", ascending=False, na_position="last").reset_index(drop=True)

    if fleet_size is None:
        fleet_size = df["vehiculos_involucrados"].max()
        if pd.isna(fleet_size) or fleet_size == 0:
            fleet_size = 1

    def clasificar(row):
        if pd.isna(row["eventos_disparados"]):
            return "ERROR AL PROCESAR"
        if row["eventos_disparados"] == 0:
            return "SIN EVENTOS - revisar lógica"
        if row["vehiculos_involucrados"] >= 0.9 * fleet_size:
            return "TODA LA FLOTA - revisar umbral"
        return "Normal"

    df["estado"] = df.apply(clasificar, axis=1)

    def formatear_top5(txt):
        if not isinstance(txt, str) or not txt.strip():
            return ""
        partes = [p.strip() for p in txt.split(";") if p.strip()]
        return "\n".join(partes)

    df["top_5_formateado"] = df["top_5_vehiculos_por_eventos"].apply(formatear_top5)

    wb = Workbook()
    FUENTE = "Arial"

    # ---------------- Hoja Resumen ----------------
    ws_r = wb.active
    ws_r.title = "Resumen"

    titulo_font = Font(name=FUENTE, size=16, bold=True, color="1F4E78")
    subtitulo_font = Font(name=FUENTE, size=10, italic=True, color="595959")
    label_font = Font(name=FUENTE, size=11, bold=True)
    header_font = Font(name=FUENTE, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    ws_r["B2"] = "Validación de Reglas Custom Geotab"
    ws_r["B2"].font = titulo_font
    ws_r["B3"] = f"Periodo analizado: {fecha_inicio_str} a {fecha_fin_str}  |  Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_r["B3"].font = subtitulo_font

    labels = [
        ("Total de reglas evaluadas", "=COUNTA(Detalle!A3:A10000)"),
        ("Reglas SIN eventos en el periodo (revisar lógica)", '=COUNTIF(Detalle!C3:C10000,"SIN EVENTOS - revisar lógica")'),
        ("Reglas disparando en casi toda la flota (revisar umbral)", '=COUNTIF(Detalle!C3:C10000,"TODA LA FLOTA - revisar umbral")'),
        ("Reglas con comportamiento normal", '=COUNTIF(Detalle!C3:C10000,"Normal")'),
        ("Reglas con error al consultar la API", '=COUNTIF(Detalle!C3:C10000,"ERROR AL PROCESAR")'),
        ("Total de eventos disparados en el periodo (todas las reglas)", "=SUM(Detalle!D3:D10000)"),
    ]

    row = 5
    for label, formula in labels:
        ws_r.cell(row=row, column=2, value=label).font = label_font
        c = ws_r.cell(row=row, column=5, value=formula)
        c.font = Font(name=FUENTE, size=11, bold=True, color="1F4E78")
        c.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    ws_r.cell(row=row, column=2, value='Cómo leer el semáforo de la hoja "Detalle":').font = label_font
    row += 1
    notas = [
        ("SIN EVENTOS - revisar lógica", "FFC7CE", "La regla no disparó ni un solo evento en el periodo. Puede tener un error de construcción (ej. una comparación que nunca se cumple) o un umbral inalcanzable."),
        ("TODA LA FLOTA - revisar umbral", "FFEB9C", "La regla disparó en el 90%+ de la flota. Puede ser comportamiento operativo normal mal marcado como excepción, o un umbral demasiado sensible."),
        ("Normal", "C6E0B4", "Volumen y alcance dentro de lo esperado; no es prioridad de revisión."),
        ("ERROR AL PROCESAR", "D9D9D9", "La consulta a la API falló para esta regla incluso después de reintentar. Revisar manualmente o volver a correr el script."),
    ]
    for texto, color, explicacion in notas:
        c = ws_r.cell(row=row, column=2, value=texto)
        c.font = Font(name=FUENTE, size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=color)
        ws_r.cell(row=row, column=3, value=explicacion).font = Font(name=FUENTE, size=10)
        ws_r.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        row += 1

    ws_r.column_dimensions["A"].width = 2
    ws_r.column_dimensions["B"].width = 46
    ws_r.column_dimensions["C"].width = 14
    ws_r.column_dimensions["D"].width = 14
    ws_r.column_dimensions["E"].width = 16

    # ---------------- Hoja Detalle ----------------
    ws = wb.create_sheet("Detalle")

    headers = [
        "Regla", "Rule ID", "Estado", "Eventos disparados", "Vehículos involucrados",
        "Diagnósticos usados", "Gap mediano muestreo (s)", "Top 5 vehículos (placa:eventos)",
    ]
    ws.append([])
    ws.append(headers)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    fill_rojo = PatternFill("solid", fgColor="FFC7CE")
    fill_amarillo = PatternFill("solid", fgColor="FFEB9C")
    fill_verde = PatternFill("solid", fgColor="C6E0B4")
    fill_gris = PatternFill("solid", fgColor="D9D9D9")

    for i, r in df.iterrows():
        fila = 3 + i
        valores = [
            r["regla"], r["rule_id"], r["estado"],
            r["eventos_disparados"] if pd.notna(r["eventos_disparados"]) else None,
            r["vehiculos_involucrados"] if pd.notna(r["vehiculos_involucrados"]) else None,
            r["diagnosticos_usados"] if pd.notna(r["diagnosticos_usados"]) else None,
            r["gap_mediano_muestreo_seg"] if pd.notna(r["gap_mediano_muestreo_seg"]) else None,
            r["top_5_formateado"],
        ]
        for col_idx, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col_idx, value=val)
            c.font = Font(name=FUENTE, size=10)
            c.border = border
            if col_idx == 8:
                c.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_idx == 1:
                c.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

        estado = r["estado"]
        if estado == "SIN EVENTOS - revisar lógica":
            fill = fill_rojo
        elif estado == "TODA LA FLOTA - revisar umbral":
            fill = fill_amarillo
        elif estado == "ERROR AL PROCESAR":
            fill = fill_gris
        else:
            fill = fill_verde
        ws.cell(row=fila, column=3).fill = fill

        n_lineas = max(1, r["top_5_formateado"].count("\n") + 1) if r["top_5_formateado"] else 1
        ws.row_dimensions[fila].height = max(15, 13 * n_lineas)

    widths = [42, 24, 26, 16, 16, 16, 18, 30]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = "C3"

    ultima_fila = 2 + len(df)
    tabla = Table(displayName="TablaValidacion", ref=f"A2:H{ultima_fila}")
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(tabla)

    wb.save(xlsx_path)
    return df["estado"].value_counts().to_dict()


# ---------------------------------------------------------------------------
# Diagnóstico de reglas en cero eventos (descarta causas triviales antes de
# asumir un error de lógica en la condición)
# ---------------------------------------------------------------------------

def _condicion_a_texto(condition, nivel=0):
    """Convierte el árbol de condiciones crudo de la API en texto indentado legible,
    mostrando tipo de condición, valor, unidad y diagnóstico asociado en cada nodo."""
    if not condition:
        return "  " * nivel + "(sin condición)\n"

    tipo = condition.get("conditionType", "?")
    partes = [tipo]
    if condition.get("value") is not None:
        partes.append(f"value={condition['value']}")
    if condition.get("unit"):
        partes.append(f"unit={condition['unit']}")
    diag = condition.get("diagnostic")
    if diag and diag.get("id"):
        partes.append(f"diagnostic={diag['id']}")

    texto = "  " * nivel + " ".join(partes) + "\n"
    for hijo in (condition.get("children") or []):
        texto += _condicion_a_texto(hijo, nivel + 1)
    return texto


def diagnosticar_reglas_cero(api, csv_previo):
    """Lee un CSV generado previamente por este script, toma las reglas con
    eventos_disparados == 0, y para cada una revisa: si está activa (state),
    si el rango activeFrom/activeTo cubre hoy, a qué grupos de vehículos está
    restringida, y muestra el árbol de condiciones crudo."""
    df = pd.read_csv(csv_previo, encoding="utf-8-sig")
    df_cero = df[df["eventos_disparados"] == 0]
    if df_cero.empty:
        print("No hay reglas con 0 eventos en ese CSV.")
        return

    print(f"Diagnosticando {len(df_cero)} regla(s) con 0 eventos...\n")
    ahora = datetime.utcnow()

    resumen_causas = defaultdict(int)

    for _, fila in df_cero.iterrows():
        rule_id = fila["rule_id"]
        nombre = fila["regla"]
        detalle = api.get("Rule", search={"id": rule_id})
        if not detalle:
            print(f"--- {nombre} --- (no se pudo recuperar el detalle de la regla)\n")
            continue
        r = detalle[0]

        print(f"--- {nombre} ---")
        causas = []

        estado = r.get("state", "")
        if "Inactive" in estado or "inactive" in estado.lower():
            causas.append("REGLA INACTIVA/DESHABILITADA")

        try:
            activo_desde = pd.to_datetime(r.get("activeFrom")).tz_localize(None) if r.get("activeFrom") else None
            activo_hasta = pd.to_datetime(r.get("activeTo")).tz_localize(None) if r.get("activeTo") else None
        except Exception:
            activo_desde = activo_hasta = None

        if activo_desde is not None and activo_desde > pd.Timestamp(ahora):
            causas.append(f"AÚN NO ENTRA EN VIGENCIA (activeFrom={activo_desde.date()})")
        if activo_hasta is not None and activo_hasta < pd.Timestamp(ahora):
            causas.append(f"YA EXPIRÓ (activeTo={activo_hasta.date()})")

        grupos = [g.get("id") for g in (r.get("groups") or [])]
        if grupos and grupos != ["GroupCompanyId"]:
            grupos_vacios = []
            for gid in grupos:
                if gid == "GroupCompanyId":
                    continue
                try:
                    dispositivos = api.get("Device", search={"groups": [{"id": gid}]})
                except Exception:
                    dispositivos = None
                n_disp = len(dispositivos) if dispositivos is not None else None
                if n_disp == 0:
                    grupos_vacios.append(gid)
            if grupos_vacios:
                causas.append(f"GRUPO SIN VEHÍCULOS ASIGNADOS (huérfana): {grupos_vacios}")
            else:
                causas.append(f"Restringida a grupo(s) con vehículos asignados (alcance limitado, no necesariamente un error): {grupos}")

        if not causas:
            causas.append("Sin causa trivial detectada -> revisar lógica de la condición")

        for c in causas:
            if c.startswith("GRUPO SIN VEHÍCULOS"):
                resumen_causas["GRUPO SIN VEHÍCULOS ASIGNADOS (huérfana)"] += 1
            elif c.startswith("Restringida a grupo"):
                resumen_causas["Restringida a grupo con vehículos (alcance normal)"] += 1
            else:
                resumen_causas[c.split(" (")[0].split(":")[0]] += 1
            print(f"  * {c}")

        print(f"  Estado: {estado} | Vigencia: {r.get('activeFrom')} a {r.get('activeTo')} | Grupos: {grupos}")
        print("  Árbol de condiciones:")
        print(_condicion_a_texto(r.get("condition"), nivel=2))

    print("=" * 70)
    print("Resumen de causas encontradas:")
    for causa, n in sorted(resumen_causas.items(), key=lambda x: -x[1]):
        print(f"  {n:>3}  {causa}")


# ---------------------------------------------------------------------------
# Corrección de reglas vía API (con respaldo y modo simulación por defecto)
# ---------------------------------------------------------------------------

def _nodo_referencia_diagnostico(condition, diagnostic_id):
    """True si este nodo tiene un hijo directo de tipo FilterStatusDataByDiagnostic
    apuntando al diagnostic_id dado (la forma real en que Geotab asocia un valor con
    un diagnóstico: el valor vive en el nodo padre, ej. IsValueLessThan, y el
    diagnóstico vive en un hijo separado, FilterStatusDataByDiagnostic)."""
    for hijo in (condition.get("children") or []):
        if hijo.get("conditionType") == "FilterStatusDataByDiagnostic":
            diag = hijo.get("diagnostic")
            if diag and diag.get("id") == diagnostic_id:
                return True
        # también soporta el caso (menos común) en que el diagnóstico esté directo en el nodo
        diag_directo = hijo.get("diagnostic")
        if hijo.get("conditionType") is None and diag_directo and diag_directo.get("id") == diagnostic_id:
            return True
    return False


def _buscar_y_modificar_nodo(condition, diagnostic_id, valor_actual, valor_nuevo, cambios_hechos):
    """Recorre el árbol de condiciones y, si encuentra un nodo con 'value' cuyo hijo
    FilterStatusDataByDiagnostic apunta al diagnostic_id esperado y el value coincide
    con valor_actual, lo reemplaza por valor_nuevo. Modifica in-place."""
    if not condition:
        return

    if condition.get("value") is not None and _nodo_referencia_diagnostico(condition, diagnostic_id):
        try:
            coincide_valor = float(condition["value"]) == float(valor_actual)
        except (TypeError, ValueError):
            coincide_valor = False
        if coincide_valor:
            cambios_hechos.append((condition["value"], valor_nuevo))
            condition["value"] = valor_nuevo

    for hijo in (condition.get("children") or []):
        _buscar_y_modificar_nodo(hijo, diagnostic_id, valor_actual, valor_nuevo, cambios_hechos)


def aplicar_correcciones(api, csv_correcciones, aplicar=False, carpeta_backup="backups_reglas"):
    """Lee un CSV con columnas: regla, diagnostic_id, valor_actual, valor_nuevo
    Para cada fila: busca la regla por nombre, localiza el nodo con ese diagnostic_id y
    valor_actual, y (si aplicar=True) lo actualiza vía api.set('Rule', ...), respaldando
    antes el objeto original completo en un JSON por regla."""
    import json
    correcciones = pd.read_csv(csv_correcciones, encoding="utf-8-sig")
    columnas_esperadas = {"regla", "diagnostic_id", "valor_actual", "valor_nuevo"}
    if not columnas_esperadas.issubset(correcciones.columns):
        print(f"El CSV debe tener las columnas: {columnas_esperadas}")
        return

    if aplicar:
        os.makedirs(carpeta_backup, exist_ok=True)

    todas_reglas = {r.get("name"): r for r in api.get("Rule")}

    for _, fila in correcciones.iterrows():
        nombre = str(fila["regla"]).strip()
        diag_id = str(fila["diagnostic_id"]).strip()
        valor_actual = fila["valor_actual"]
        valor_nuevo = fila["valor_nuevo"]

        print(f"--- {nombre} ---")
        rule = todas_reglas.get(nombre)
        if rule is None:
            print(f"  *** No se encontró una regla con el nombre exacto '{nombre}'. Sáltala o corrige el nombre en el CSV. ***\n")
            continue

        # Traer la copia completa y actualizada de la regla (por si el listado inicial está desfasado)
        detalle = api.get("Rule", search={"id": rule["id"]})
        if not detalle:
            print("  *** No se pudo recuperar el detalle de la regla. ***\n")
            continue
        regla_completa = detalle[0]

        # IMPORTANTE: capturar el respaldo ANTES de modificar nada en memoria
        respaldo_original = json.loads(json.dumps(regla_completa, default=str))

        cambios_hechos = []
        _buscar_y_modificar_nodo(regla_completa.get("condition"), diag_id, valor_actual, valor_nuevo, cambios_hechos)

        if not cambios_hechos:
            print(f"  *** No se encontró ningún nodo con diagnostic_id={diag_id} y valor_actual={valor_actual}. "
                  f"Nada que cambiar (verifica los datos del CSV contra el árbol real de la regla). ***\n")
            continue

        for viejo, nuevo in cambios_hechos:
            print(f"  Cambio detectado: {viejo}  ->  {nuevo}")

        if not aplicar:
            print("  (modo simulación: no se guardó nada. Corre con --apply para aplicar el cambio de verdad)\n")
            continue

        # Respaldo del objeto original (capturado antes de la modificación)
        backup_path = os.path.join(
            carpeta_backup,
            f"{nombre.replace(' ', '_').replace('/', '-')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        with open(backup_path, "w", encoding="utf-8") as f:
            json.dump(respaldo_original, f, ensure_ascii=False, indent=2, default=str)
        print(f"  Respaldo guardado en: {backup_path}")

        try:
            api.set("Rule", regla_completa)
            print("  *** Cambio aplicado correctamente en MyGeotab. ***\n")
        except Exception as e:
            print(f"  *** ERROR al guardar el cambio: {e}. La regla NO quedó modificada (o revisa manualmente). ***\n")


# ---------------------------------------------------------------------------
# Inspección de valores reales de un diagnóstico (para calibrar umbrales con
# datos reales de la flota antes de corregir una regla, en vez de adivinar)
# ---------------------------------------------------------------------------

def inspeccionar_valores_diagnostico(api, diagnostic_id, fecha_inicio, fecha_fin, devices=None, muestras_por_device=5):
    """Trae una muestra de StatusData cruda para un diagnóstico específico (en varios
    vehículos si no se especifica una lista) y muestra min/max/mediana + algunos valores
    de ejemplo, para confirmar la escala real en la que Geotab está capturando ese dato."""
    if not devices:
        todos = api.get("Device")
        devices = [d["id"] for d in todos[:15]]  # muestra de hasta 15 vehículos si no se especifica

    print(f"Inspeccionando diagnóstico {diagnostic_id} en {len(devices)} vehículo(s)...\n")

    todos_los_valores = []
    for device_id in devices:
        datos = contar_statusdata(api, device_id, diagnostic_id, fecha_inicio, fecha_fin)
        if not datos:
            continue
        valores = [d["data"] for d in datos if d.get("data") is not None]
        if not valores:
            continue
        muestra = valores[:muestras_por_device]
        print(f"  Dispositivo {device_id}: {len(valores)} muestras. "
              f"min={min(valores):.2f}  max={max(valores):.2f}  ejemplo={muestra}")
        todos_los_valores.extend(valores)

    if not todos_los_valores:
        print("\n*** No se encontraron datos para este diagnóstico en el periodo/vehículos indicados. ***")
        print("    Esto en sí mismo es información: puede confirmar que el diagnóstico nunca llega como StatusData.")
        return

    s = pd.Series(todos_los_valores)
    print("\n" + "=" * 60)
    print(f"RESUMEN GLOBAL ({len(todos_los_valores)} muestras de {len(devices)} vehículos):")
    print(f"  Mínimo:  {s.min():,.2f}")
    print(f"  Máximo:  {s.max():,.2f}")
    print(f"  Mediana: {s.median():,.2f}")
    print(f"  Percentil 5:  {s.quantile(0.05):,.2f}")
    print(f"  Percentil 95: {s.quantile(0.95):,.2f}")


# ---------------------------------------------------------------------------
# Creación de una regla nueva vía API (con modo simulación por defecto)
# ---------------------------------------------------------------------------

def crear_regla(api, nombre, condition, grupos=None, aplicar=False):
    """Crea una nueva regla en MyGeotab vía api.add('Rule', ...). Por defecto corre en
    modo simulación (solo muestra el árbol que se crearía). Con aplicar=True, la crea
    de verdad y devuelve el id nuevo."""
    ahora = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")
    entidad = {
        "name": nombre,
        "condition": condition,
        "groups": grupos or [{"id": "GroupCompanyId"}],
        "state": "ExceptionRuleStateActiveId",
        "reason": "ExceptionRuleReasonNoneId",
        "comment": "",
        "activeFrom": ahora,
        "activeTo": "2050-01-01T00:00:00.000Z",
        "color": {"a": 255, "r": 255, "g": 165, "b": 0},
    }

    print(f"--- Regla a crear: {nombre} ---")
    print(_condicion_a_texto(condition))
    print(f"Grupos: {entidad['groups']}")

    if not aplicar:
        print("(modo simulación: no se creó nada. Corre con --apply para crearla de verdad)")
        return None

    try:
        nuevo_id = api.add("Rule", entidad)
        print(f"*** Regla creada correctamente en MyGeotab. id={nuevo_id} ***")
        return nuevo_id
    except Exception as e:
        print(f"*** ERROR al crear la regla: {e} ***")
        print("    Si el error persiste, probablemente Geotab exige otro campo no documentado en el objeto Rule.")
        print("    Alternativa confiable: crea la regla manualmente en MyGeotab (Reglas y grupos > Reglas > Añadir),")
        print("    usando exactamente el árbol de condiciones impreso arriba como guía.")
        return None


def construir_condicion_ralenti_excesivo(duracion_seg=300, velocidad_max_kmh=1, pto_diagnostic_id="DiagnosticPowerTakeoffEngagedId"):
    """Árbol de condición para 'ralentí excesivo, sin actividad de PTO':
    Ignition=true AND Speed < velocidad_max_kmh AND PTO no activo, sostenido >= duracion_seg."""
    return {
        "conditionType": "DurationLongerThan", "value": duracion_seg, "unit": "s",
        "children": [
            {
                "conditionType": "And",
                "children": [
                    {"conditionType": "Ignition", "value": 1},
                    {
                        "conditionType": "IsValueLessThan", "value": velocidad_max_kmh, "unit": "km/h",
                        "children": [{"conditionType": "Speed"}],
                    },
                    {
                        "conditionType": "IsValueLessThan", "value": 1,
                        "children": [{"conditionType": "FilterStatusDataByDiagnostic",
                                       "diagnostic": {"id": pto_diagnostic_id}}],
                    },
                ],
            }
        ],
    }


def _quitar_ids_recursivo(condition):
    """Quita el campo 'id' de cada nodo del árbol de condiciones (y 'sequence'), para que al
    clonar una regla existente Geotab genere identificadores nuevos en vez de chocar con los
    de la regla original."""
    if not condition:
        return
    condition.pop("id", None)
    condition.pop("sequence", None)
    for hijo in (condition.get("children") or []):
        _quitar_ids_recursivo(hijo)


def crear_regla_clonando(api, nombre_regla_template, nombre_nuevo, transformar_condicion, aplicar=False):
    """Alternativa más robusta a crear_regla: en vez de construir el objeto Rule desde cero
    (arriesgando que falte algún campo no documentado), clona una regla EXISTENTE que ya
    funciona, le aplica una transformación a su árbol de condiciones, y la guarda como nueva.
    transformar_condicion: función que recibe el árbol de condiciones original (dict) y
    devuelve el árbol modificado."""
    todas = api.get("Rule")
    template = next((r for r in todas if r.get("name") == nombre_regla_template), None)
    if template is None:
        print(f"*** No se encontró una regla llamada exactamente '{nombre_regla_template}' para usar como plantilla. ***")
        return None

    detalle = api.get("Rule", search={"id": template["id"]})
    if not detalle:
        print("*** No se pudo recuperar el detalle de la regla plantilla. ***")
        return None

    import copy
    nueva_entidad = copy.deepcopy(detalle[0])
    nueva_entidad.pop("id", None)
    nueva_entidad.pop("version", None)
    nueva_entidad["name"] = nombre_nuevo

    nueva_entidad["condition"] = transformar_condicion(nueva_entidad["condition"])
    _quitar_ids_recursivo(nueva_entidad["condition"])

    print(f"--- Regla a crear (clonada de '{nombre_regla_template}'): {nombre_nuevo} ---")
    print(_condicion_a_texto(nueva_entidad["condition"]))

    if not aplicar:
        print("(modo simulación: no se creó nada. Corre con --apply para crearla de verdad)")
        return None

    try:
        nuevo_id = api.add("Rule", nueva_entidad)
        print(f"*** Regla creada correctamente en MyGeotab. id={nuevo_id} ***")
        return nuevo_id
    except Exception as e:
        print(f"*** ERROR al crear la regla clonada: {e} ***")
        return None


# ---------------------------------------------------------------------------
# Cambiar el/los grupo(s) de vehículos a los que aplica una regla existente
# ---------------------------------------------------------------------------

def actualizar_grupos_regla(api, nombre_regla, grupos_nuevos, aplicar=False, carpeta_backup="backups_reglas"):
    """Cambia el campo 'groups' de una regla existente (ej. para que aplique a toda la
    flota en vez de a un grupo específico heredado de una plantilla clonada).
    grupos_nuevos: lista de group ids, ej. ['GroupCompanyId'] para toda la flota."""
    import json
    todas = api.get("Rule")
    rule = next((r for r in todas if r.get("name") == nombre_regla), None)
    if rule is None:
        print(f"*** No se encontró una regla llamada exactamente '{nombre_regla}'. ***")
        return

    detalle = api.get("Rule", search={"id": rule["id"]})
    if not detalle:
        print("*** No se pudo recuperar el detalle de la regla. ***")
        return
    regla_completa = detalle[0]

    grupos_actuales = [g.get("id") for g in (regla_completa.get("groups") or [])]
    print(f"--- {nombre_regla} ---")
    print(f"  Grupos actuales: {grupos_actuales}")
    print(f"  Grupos nuevos:   {grupos_nuevos}")

    if grupos_actuales == grupos_nuevos:
        print("  Ya están iguales, no hay nada que cambiar.\n")
        return

    if not aplicar:
        print("  (modo simulación: no se guardó nada. Corre con --apply para aplicar el cambio de verdad)\n")
        return

    respaldo_original = json.loads(json.dumps(regla_completa, default=str))
    os.makedirs(carpeta_backup, exist_ok=True)
    backup_path = os.path.join(
        carpeta_backup,
        f"{nombre_regla.replace(' ', '_').replace('/', '-')}_grupos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    )
    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(respaldo_original, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Respaldo guardado en: {backup_path}")

    regla_completa["groups"] = [{"id": g} for g in grupos_nuevos]
    try:
        api.set("Rule", regla_completa)
        print("  *** Grupos actualizados correctamente en MyGeotab. ***\n")
    except Exception as e:
        print(f"  *** ERROR al guardar el cambio: {e} ***\n")


def detalle_eventos_notables(eventos, devices=None, min_duracion_min=20):
    """Imprime el detalle exacto (vehículo, fecha/hora de inicio, duración) de los eventos
    que correspondan a los device_id indicados en 'devices', y/o que superen min_duracion_min,
    ordenados de mayor a menor duración. Útil para cruzar contra el mapa/historial de viajes."""
    filas = []
    for e in eventos:
        dur_min = _parsear_duracion_a_minutos(e.get("duration"))
        if dur_min is None:
            continue
        dev_id = (e.get("device") or {}).get("id", "desconocido")
        incluir = False
        if devices and dev_id in devices:
            incluir = True
        if min_duracion_min is not None and dur_min >= min_duracion_min:
            incluir = True
        if incluir:
            filas.append({
                "device_id": placa_de(dev_id),
                "inicio": _normalizar_fecha_para_excel(e.get("activeFrom")),
                "duracion_min": round(dur_min, 1),
            })

    if not filas:
        print("No hay eventos que coincidan con los filtros de detalle.")
        return

    filas.sort(key=lambda f: f["duracion_min"], reverse=True)
    print(f"\n=== DETALLE DE EVENTOS NOTABLES ({len(filas)}) — hora Bogotá ===")
    for f in filas:
        print(f"  {f['device_id']:>8}  |  inicio: {f['inicio']}  |  duración: {f['duracion_min']} min")


# ---------------------------------------------------------------------------
# Ajustar la duración mínima (DurationLongerThan) de una regla existente
# ---------------------------------------------------------------------------

def _buscar_nodo_por_tipo(condition, tipo):
    """Devuelve (por referencia) el primer nodo del árbol cuyo conditionType coincide."""
    if not condition:
        return None
    if condition.get("conditionType") == tipo:
        return condition
    for hijo in (condition.get("children") or []):
        encontrado = _buscar_nodo_por_tipo(hijo, tipo)
        if encontrado is not None:
            return encontrado
    return None


def actualizar_duracion_regla(api, nombre_regla, nuevo_valor_seg, aplicar=False, carpeta_backup="backups_reglas"):
    """Cambia el 'value' del primer nodo DurationLongerThan de una regla existente."""
    import json
    todas = api.get("Rule")
    rule = next((r for r in todas if r.get("name") == nombre_regla), None)
    if rule is None:
        print(f"*** No se encontró una regla llamada exactamente '{nombre_regla}'. ***")
        return

    detalle = api.get("Rule", search={"id": rule["id"]})
    if not detalle:
        print("*** No se pudo recuperar el detalle de la regla. ***")
        return
    regla_completa = detalle[0]

    respaldo_original = json.loads(json.dumps(regla_completa, default=str))

    nodo = _buscar_nodo_por_tipo(regla_completa.get("condition"), "DurationLongerThan")
    if nodo is None:
        print(f"*** Esta regla no tiene ningún nodo 'DurationLongerThan' que ajustar. ***")
        return

    valor_actual = nodo.get("value")
    print(f"--- {nombre_regla} ---")
    print(f"  Duración actual: {valor_actual} {nodo.get('unit', 's')}")
    print(f"  Duración nueva:  {nuevo_valor_seg} s ({nuevo_valor_seg/60:.1f} min)")

    if valor_actual == nuevo_valor_seg:
        print("  Ya está en ese valor, no hay nada que cambiar.\n")
        return

    if not aplicar:
        print("  (modo simulación: no se guardó nada. Corre con --apply para aplicar el cambio de verdad)\n")
        return

    os.makedirs(carpeta_backup, exist_ok=True)
    backup_path = os.path.join(
        carpeta_backup,
        f"{nombre_regla.replace(' ', '_').replace('/', '-')}_duracion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    )
    with open(backup_path, "w", encoding="utf-8") as f:
        json.dump(respaldo_original, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Respaldo guardado en: {backup_path}")

    nodo["value"] = nuevo_valor_seg
    try:
        api.set("Rule", regla_completa)
        print("  *** Duración actualizada correctamente en MyGeotab. ***\n")
    except Exception as e:
        print(f"  *** ERROR al guardar el cambio: {e} ***")
        print("      Si vuelve a dar GenericException, aplica este cambio manualmente en MyGeotab")
        print("      (mismo problema que vimos al cambiar el grupo por API).\n")


# ---------------------------------------------------------------------------
# Traducción de device_id (código interno de Geotab) a placa real del vehículo
# ---------------------------------------------------------------------------

_MAPA_PLACAS = {}


def cargar_mapa_placas(api):
    """Trae todos los Device y arma un diccionario device_id -> placa, para mostrar
    placas reales en los reportes en vez del código interno (ej. 'b55C')."""
    global _MAPA_PLACAS
    try:
        devices = api.get("Device")
    except Exception as e:
        print(f"(no se pudo cargar el mapa de placas: {e}, se usarán los IDs internos)")
        devices = []

    mapa = {}
    for d in devices:
        placa = (d.get("licensePlate") or "").strip()
        nombre = (d.get("name") or "").strip()
        mapa[d["id"]] = placa or nombre or d["id"]
    _MAPA_PLACAS = mapa
    return mapa


def placa_de(device_id):
    """Devuelve la placa real de un device_id, o el mismo id si no se encontró."""
    return _MAPA_PLACAS.get(device_id, device_id)


# ---------------------------------------------------------------------------
# Ubicar dónde estaba un vehículo en un momento exacto (para investigar eventos atípicos)
# ---------------------------------------------------------------------------

def _placa_a_device_id(placa):
    """Búsqueda inversa: dada una placa, devuelve el device_id interno de Geotab."""
    for dev_id, p in _MAPA_PLACAS.items():
        if p == placa:
            return dev_id
    return None


def ubicar_evento(api, placa_o_device_id, fecha_hora_bogota_str, ventana_min=10):
    """Dado un vehículo (placa o device_id) y una fecha/hora en horario de Bogotá
    (formato 'YYYY-MM-DD HH:MM:SS'), busca el LogRecord (posición GPS) más cercano a
    ese momento y muestra las coordenadas + un enlace directo a Google Maps."""
    device_id = placa_o_device_id
    if placa_o_device_id in _MAPA_PLACAS.values():
        device_id = _placa_a_device_id(placa_o_device_id)
    if device_id is None:
        print(f"*** No se encontró el vehículo '{placa_o_device_id}' (ni como placa ni como device_id). ***")
        return

    momento_bogota = datetime.strptime(fecha_hora_bogota_str, "%Y-%m-%d %H:%M:%S")
    momento_utc = momento_bogota + timedelta(hours=5)  # Bogotá (UTC-5) -> UTC
    desde = momento_utc - timedelta(minutes=ventana_min)
    hasta = momento_utc + timedelta(minutes=ventana_min)

    search = {
        "deviceSearch": {"id": device_id},
        "fromDate": desde.isoformat(),
        "toDate": hasta.isoformat(),
    }
    registros = _get_con_reintento(api, "LogRecord", search)
    if not registros:
        print(f"*** No se encontraron posiciones GPS para {placa_de(device_id)} cerca de {fecha_hora_bogota_str} (hora Bogotá). ***")
        return

    # Encontrar el registro más cercano al momento exacto solicitado
    def _dist(r):
        try:
            t = pd.to_datetime(r["dateTime"]).tz_localize(None)
            return abs((t - momento_utc).total_seconds())
        except Exception:
            return float("inf")

    mas_cercano = min(registros, key=_dist)
    lat, lon = mas_cercano.get("latitude"), mas_cercano.get("longitude")

    print(f"\n=== Ubicación de {placa_de(device_id)} cerca de {fecha_hora_bogota_str} (hora Bogotá) ===")
    print(f"  Registro GPS más cercano: {mas_cercano.get('dateTime')} (UTC)")
    print(f"  Coordenadas: {lat}, {lon}")
    print(f"  Ver en Google Maps: https://www.google.com/maps?q={lat},{lon}")
    print(f"  ({len(registros)} posiciones GPS encontradas en la ventana de ±{ventana_min} min)")


def rango_bogota_a_utc(start_str, end_str):
    """Convierte un rango de fechas dado en día calendario de Bogotá (ej. '2026-08-15' a
    '2026-08-15' = todo el 15 de agosto en hora Colombia) al rango UTC equivalente que hay
    que enviarle a la API de Geotab. Bogotá es UTC-5 todo el año (sin horario de verano),
    así que la medianoche de Bogotá corresponde a las 05:00 UTC del mismo día."""
    inicio_bogota = datetime.strptime(start_str, "%Y-%m-%d")
    fin_bogota = datetime.strptime(end_str, "%Y-%m-%d") + timedelta(days=1)
    fecha_inicio_utc = inicio_bogota + timedelta(hours=5)
    fecha_fin_utc = fin_bogota + timedelta(hours=5)
    return fecha_inicio_utc, fecha_fin_utc


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs("reportes", exist_ok=True)  # no va a git (.gitignore), puede no existir aun

    parser = argparse.ArgumentParser(description="Valida reglas Geotab contra datos reales de telemetría")
    parser.add_argument("--start", default=None, help="Fecha inicio YYYY-MM-DD (requerido salvo con --apply-corrections)")
    parser.add_argument("--end", default=None, help="Fecha fin YYYY-MM-DD (requerido salvo con --apply-corrections)")
    parser.add_argument("--rules", nargs="*", default=[],
                         help="Nombre (o parte del nombre) de cada regla a validar. Si se omite, valida TODAS las reglas de la base de datos.")
    parser.add_argument("--devices", nargs="*", default=[],
                         help="(Opcional) Device IDs puntuales para revisar el gap de muestreo real de sus diagnósticos.")
    parser.add_argument("--out", default="reportes/validacion_reglas.csv", help="Archivo CSV de salida")
    parser.add_argument("--list", action="store_true",
                         help="Solo lista los nombres de TODAS las reglas existentes en la base de datos (no valida nada) y termina.")
    parser.add_argument("--inspect", default=None,
                         help="Nombre exacto (o parte del nombre) de UNA regla para imprimir todos sus campos crudos desde la API (para descubrir cómo distinguir reglas custom vs predefinidas) y termina.")
    parser.add_argument("--custom-only", action="store_true",
                         help="Excluye las reglas predefinidas de Geotab (baseType == 'Stock') y deja solo las creadas por el equipo.")
    parser.add_argument("--fleet-size", type=int, default=None,
                         help="(Opcional) Tamaño de flota a usar para el umbral 'TODA LA FLOTA' del Excel. Por defecto se autodetecta como el máximo de vehículos involucrados encontrado.")
    parser.add_argument("--no-excel", action="store_true",
                         help="No generar el Excel formateado al final, solo el CSV.")
    parser.add_argument("--diagnose-zero", default=None, metavar="CSV_PREVIO",
                         help="Ruta a un CSV generado previamente por este script. Revisa las reglas con 0 eventos "
                              "y descarta causas triviales (regla inactiva, fuera de vigencia, restringida a un "
                              "grupo) antes de mostrar el árbol de condiciones para revisión de lógica. Termina sin "
                              "correr la validación normal.")
    parser.add_argument("--apply-corrections", default=None, metavar="CSV_CORRECCIONES",
                         help="Ruta a un CSV con columnas: regla,diagnostic_id,valor_actual,valor_nuevo. "
                              "Por defecto corre en modo simulación (no cambia nada). Usa junto con --apply para "
                              "aplicar los cambios de verdad. Termina sin correr la validación normal.")
    parser.add_argument("--apply", action="store_true",
                         help="Junto con --apply-corrections: aplica los cambios de verdad en MyGeotab (guarda respaldo antes).")
    parser.add_argument("--inspect-diagnostic", default=None, metavar="DIAGNOSTIC_ID",
                         help="Consulta valores crudos reales de StatusData para un diagnostic_id específico "
                              "(muestra min/max/mediana/percentiles). Úsalo para calibrar el umbral correcto de una "
                              "regla con datos reales, antes de escribir un CSV de corrección. Termina sin correr "
                              "la validación normal. Usa --devices para limitar a vehículos específicos (si se "
                              "omite, toma una muestra de hasta 15 vehículos de la flota).")
    parser.add_argument("--create-idle-rule", default=None, metavar="NOMBRE_REGLA",
                         help="Crea (o simula crear, sin --apply) una nueva regla de 'ralentí excesivo sin PTO "
                              "activo' con el nombre dado. Usa --idle-duration-min (default 5) y "
                              "--pto-diagnostic-id (default DiagnosticPowerTakeoffEngagedId) para ajustarla. "
                              "Termina sin correr la validación normal.")
    parser.add_argument("--idle-duration-min", type=float, default=5,
                         help="(Con --create-idle-rule) Minutos mínimos de ralentí sostenido para disparar. Default: 5.")
    parser.add_argument("--pto-diagnostic-id", default="DiagnosticPowerTakeoffEngagedId",
                         help="(Con --create-idle-rule) diagnostic id del PTO a excluir. Default: DiagnosticPowerTakeoffEngagedId.")
    parser.add_argument("--clone-idle-rule", default=None, metavar="NOMBRE_REGLA_NUEVA",
                         help="Alternativa más robusta a --create-idle-rule: clona la regla existente "
                              "'ALERTA EN REVOLUCIONES ALTAS CON PTO (X12)' (o la que indiques con "
                              "--template-rule) y ajusta su condición a 'ralentí excesivo sin PTO activo', "
                              "en vez de construir el objeto desde cero. Usa esto si --create-idle-rule falla "
                              "con un error genérico de la API.")
    parser.add_argument("--template-rule", default="ALERTA EN REVOLUCIONES ALTAS CON PTO (X12)",
                         help="(Con --clone-idle-rule) Nombre exacto de la regla existente a usar como plantilla.")
    parser.add_argument("--set-rule-groups", default=None, metavar="NOMBRE_REGLA",
                         help="Cambia el grupo de vehículos de una regla existente. Usa --groups para indicar "
                              "los group id nuevos (ej. --groups GroupCompanyId para toda la flota). "
                              "Termina sin correr la validación normal.")
    parser.add_argument("--groups", nargs="*", default=["GroupCompanyId"],
                         help="(Con --set-rule-groups) Lista de group id a asignar. Default: GroupCompanyId (toda la flota).")
    parser.add_argument("--duration-report", default=None, metavar="NOMBRE_REGLA",
                         help="Descarga los eventos de una regla (usa --start/--end) y muestra la distribución "
                              "real de duración (min/percentiles/rangos) además del conteo por vehículo. Útil para "
                              "decidir si el umbral de duración de una regla está bien calibrado. Con --devices, "
                              "también imprime el detalle exacto (fecha/hora/duración) de esos vehículos. Termina "
                              "sin correr la validación normal.")
    parser.add_argument("--set-duration", default=None, metavar="NOMBRE_REGLA",
                         help="Cambia la duración mínima (nodo DurationLongerThan) de una regla existente. Usa "
                              "--duration-min para el valor nuevo en minutos. Termina sin correr la validación normal.")
    parser.add_argument("--duration-min", type=float, default=None,
                         help="(Con --set-duration) Nuevo valor de duración mínima, en minutos.")
    parser.add_argument("--locate-event", default=None, metavar="PLACA_O_DEVICE_ID",
                         help="Ubica dónde estaba un vehículo en un momento exacto (útil para investigar un evento "
                              "atípico). Usa junto con --event-time 'YYYY-MM-DD HH:MM:SS' (hora Bogotá). Muestra "
                              "coordenadas GPS y un enlace a Google Maps. Termina sin correr la validación normal.")
    parser.add_argument("--event-time", default=None, metavar="'YYYY-MM-DD HH:MM:SS'",
                         help="(Con --locate-event) Fecha y hora exacta del evento, en hora de Bogotá.")
    args = parser.parse_args()

    if args.diagnose_zero:
        api = conectar()
        diagnosticar_reglas_cero(api, args.diagnose_zero)
        sys.exit(0)

    if args.apply_corrections:
        api = conectar()
        aplicar_correcciones(api, args.apply_corrections, aplicar=args.apply)
        sys.exit(0)

    if args.locate_event:
        if not args.event_time:
            print("*** --locate-event requiere --event-time 'YYYY-MM-DD HH:MM:SS'. ***")
            sys.exit(1)
        api = conectar()
        ubicar_evento(api, args.locate_event, args.event_time)
        sys.exit(0)

    if args.set_rule_groups:
        api = conectar()
        actualizar_grupos_regla(api, args.set_rule_groups, args.groups, aplicar=args.apply)
        sys.exit(0)

    if args.set_duration:
        if args.duration_min is None:
            print("*** --set-duration requiere --duration-min (en minutos). ***")
            sys.exit(1)
        api = conectar()
        actualizar_duracion_regla(api, args.set_duration, int(args.duration_min * 60), aplicar=args.apply)
        sys.exit(0)

    if args.duration_report:
        if not args.start or not args.end:
            print("*** --duration-report requiere --start y --end. ***")
            sys.exit(1)
        api = conectar()
        fecha_inicio, fecha_fin = rango_bogota_a_utc(args.start, args.end)
        todas = api.get("Rule")
        rule = next((r for r in todas if r.get("name") == args.duration_report), None)
        if rule is None:
            print(f"*** No se encontró una regla llamada exactamente '{args.duration_report}'. ***")
            sys.exit(1)
        eventos = contar_eventos_excepcion(api, rule, fecha_inicio, fecha_fin)
        print(f"Regla: {args.duration_report} — {len(eventos)} eventos descargados en el periodo.")
        df = analizar_duracion_eventos(eventos)
        detalle_eventos_notables(eventos, devices=args.devices, min_duracion_min=20)

        if df is not None and not args.no_excel:
            nombre_archivo = "detalle_duracion_" + args.duration_report.replace(" ", "_").replace("/", "-") + ".xlsx"
            print(f"\nGenerando Excel con el detalle completo de los {len(df)} eventos en: {nombre_archivo} ...")
            try:
                generar_reporte_duracion_excel(df, args.duration_report, nombre_archivo)
                print(f"Listo: {nombre_archivo}")
            except Exception as e:
                print(f"*** No se pudo generar el Excel: {e} ***")
        sys.exit(0)

    if args.create_idle_rule:
        api = conectar()
        condition = construir_condicion_ralenti_excesivo(
            duracion_seg=int(args.idle_duration_min * 60),
            pto_diagnostic_id=args.pto_diagnostic_id,
        )
        crear_regla(api, args.create_idle_rule, condition, aplicar=args.apply)
        sys.exit(0)

    if args.clone_idle_rule:
        api = conectar()

        def _transformar(condition_original):
            # Reemplaza el árbol completo de la plantilla por nuestra condición de
            # "ralentí excesivo sin PTO", reutilizando solo el resto del objeto Rule
            # (activeFrom, activeTo, color, groups, etc.) que ya sabemos que Geotab acepta.
            return construir_condicion_ralenti_excesivo(
                duracion_seg=int(args.idle_duration_min * 60),
                pto_diagnostic_id=args.pto_diagnostic_id,
            )

        crear_regla_clonando(api, args.template_rule, args.clone_idle_rule, _transformar, aplicar=args.apply)
        sys.exit(0)

    # A partir de aquí, todos los modos restantes sí necesitan --start y --end
    if not args.start or not args.end:
        print("*** Este modo requiere --start y --end (formato YYYY-MM-DD). ***")
        sys.exit(1)

    if args.inspect_diagnostic:
        api = conectar()
        fecha_inicio, fecha_fin = rango_bogota_a_utc(args.start, args.end)
        inspeccionar_valores_diagnostico(api, args.inspect_diagnostic, fecha_inicio, fecha_fin, devices=args.devices)
        sys.exit(0)

    fecha_inicio, fecha_fin = rango_bogota_a_utc(args.start, args.end)

    api = conectar()

    if args.inspect:
        todas = api.get("Rule")
        coincidencias = [r for r in todas if args.inspect.lower() in (r.get("name") or "").lower()]
        if not coincidencias:
            print(f"No se encontró ninguna regla que contenga '{args.inspect}'.")
            sys.exit(1)
        import json
        for r in coincidencias[:3]:
            print(f"\n===== {r.get('name')} =====")
            print(json.dumps(r, indent=2, ensure_ascii=False, default=str))
        sys.exit(0)

    if args.list:
        print("Buscando reglas...\n")
        todas = api.get("Rule")
        if args.custom_only:
            todas = [r for r in todas if (r.get("baseType") or "").lower() != "stock"]
            print(f"Se encontraron {len(todas)} reglas CUSTOM (excluyendo las predefinidas de Geotab con baseType='Stock'):\n")
        else:
            print(f"Se encontraron {len(todas)} reglas en total:\n")
        for r in sorted(todas, key=lambda x: (x.get("name") or "").lower()):
            print(f"  - [{r.get('baseType')}] {r.get('name')}")
        print("\nCopia el nombre exacto (o una parte única de él) para usarlo con --rules.")
        sys.exit(0)

    print("Buscando reglas...")
    reglas = obtener_reglas(api, args.rules, solo_custom=args.custom_only)
    if not reglas:
        print("No se encontraron reglas con ese nombre. Revisa el texto exacto (o parcial) del nombre en MyGeotab, "
              "o corre el script con --list para ver todos los nombres disponibles.")
        sys.exit(1)

    print(f"Se encontraron {len(reglas)} regla(s) a validar.\n")
    filas = []

    for idx, rule in enumerate(reglas, 1):
        nombre = rule.get("name", "(sin nombre)")
        print(f"--- [{idx}/{len(reglas)}] {nombre} ---")

        try:
            eventos = contar_eventos_excepcion(api, rule, fecha_inicio, fecha_fin)
            n_eventos = len(eventos)
            conteo_dev = resumen_por_dispositivo(eventos)
            n_vehiculos = len(conteo_dev)

            print(f"  Eventos disparados en el periodo: {n_eventos}")
            print(f"  Vehículos distintos involucrados: {n_vehiculos}")

            if n_eventos == 0:
                print("  *** ALERTA: esta regla NO disparó ningún evento en el periodo. "
                      "Revisar la lógica de la condición (posible error como IsValueLessThan[value=0] "
                      "sobre un campo que nunca es negativo). ***")

            diagnosticos = extraer_diagnosticos_de_condicion(rule.get("condition"))
            print(f"  Diagnósticos usados en la condición: {len(diagnosticos)}")

            gaps_detectados = []
            if args.devices and diagnosticos:
                for device_id in args.devices:
                    for diag_id in diagnosticos:
                        datos = contar_statusdata(api, device_id, diag_id, fecha_inicio, fecha_fin)
                        gap = gap_mediano_segundos(datos)
                        gaps_detectados.append(gap)
                        print(f"    Dispositivo {device_id} / Diagnóstico {diag_id}: "
                              f"{len(datos)} muestras, gap mediano = {gap} s")

            top_dev = sorted(conteo_dev.items(), key=lambda x: -x[1])[:5]

            filas.append({
                "regla": nombre,
                "rule_id": rule.get("id"),
                "eventos_disparados": n_eventos,
                "vehiculos_involucrados": n_vehiculos,
                "top_5_vehiculos_por_eventos": "; ".join(f"{placa_de(d)}:{c}" for d, c in top_dev),
                "diagnosticos_usados": len(diagnosticos),
                "gap_mediano_muestreo_seg": (
                    round(sum(g for g in gaps_detectados if g is not None) / len([g for g in gaps_detectados if g is not None]), 1)
                    if any(g is not None for g in gaps_detectados) else None
                ),
                "error": "",
            })

        except Exception as e:
            print(f"  *** ERROR procesando esta regla, se omite y se continúa con la siguiente: {e} ***")
            filas.append({
                "regla": nombre,
                "rule_id": rule.get("id"),
                "eventos_disparados": None,
                "vehiculos_involucrados": None,
                "top_5_vehiculos_por_eventos": "",
                "diagnosticos_usados": None,
                "gap_mediano_muestreo_seg": None,
                "error": str(e),
            })

        # Guardado progresivo: si el script se cae más adelante, no se pierde lo ya procesado
        try:
            pd.DataFrame(filas).to_csv(args.out, index=False, encoding="utf-8-sig")
        except PermissionError:
            print(f"  *** No se pudo escribir '{args.out}' (¿está abierto en Excel u otro programa? Ciérralo). "
                  f"Se continúa procesando; se reintentará guardar en la siguiente regla. ***")
        print()

    try:
        pd.DataFrame(filas).to_csv(args.out, index=False, encoding="utf-8-sig")
        print(f"Resumen final guardado en: {args.out}")
    except PermissionError:
        print(f"*** No se pudo guardar el CSV final en '{args.out}' porque está abierto en otro programa. "
              f"Ciérralo y corre de nuevo el script (o copia los datos que ya se imprimieron en consola). ***")

    if not args.no_excel:
        xlsx_path = args.out.rsplit(".", 1)[0] + ".xlsx"
        print(f"Generando reporte Excel formateado en: {xlsx_path} ...")
        try:
            conteos = generar_reporte_excel(args.out, xlsx_path, args.start, args.end, fleet_size=args.fleet_size)
            print(f"Reporte Excel generado correctamente: {xlsx_path}")
            print(f"  Resumen por estado: {conteos}")
        except PermissionError:
            print(f"*** No se pudo escribir '{xlsx_path}' porque está abierto en Excel. Ciérralo y vuelve a correr el script. ***")
        except Exception as e:
            print(f"*** No se pudo generar el Excel automáticamente: {e} ***")
            print(f"    El CSV sigue disponible en: {args.out}")


if __name__ == "__main__":
    main()