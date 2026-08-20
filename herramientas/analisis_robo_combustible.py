"""
Detección de posibles robos de combustible.

Busca caídas bruscas del nivel de combustible (%) en ventanas cortas de tiempo,
por vehículo, para identificar si el patrón se repite en el mismo móvil o
en varios distintos.

Requiere las mismas credenciales de Geotab que ya usa el tablero (variables
de entorno GEOTAB_USUARIO, GEOTAB_CONTRASENA, GEOTAB_DATABASE, GEOTAB_SERVER,
o un archivo .env en la misma carpeta).
"""

import os
import time
import datetime
import mygeotab
import pandas as pd
import requests
from dotenv import load_dotenv

load_dotenv()  # carga las variables desde el archivo .env en la misma carpeta

ZONAS_SOSPECHOSAS = ['GUERRERO', 'SEGUIMIENTO']  # agrega aquí más nombres para filtrar por varios barrios/zonas

MAX_REINTENTOS = 4          # cuántas veces reintentar una llamada a la API si el servidor falla
ESPERA_BASE_SEGUNDOS = 8    # espera antes del primer reintento; se duplica en cada intento siguiente


def _es_error_transitorio(error):
    """Identifica errores que valen la pena reintentar: caídas temporales del servidor
    (502/503/504), timeouts o problemas de conexión — muy comunes en el servidor del
    proveedor de telemetría (Copiloto) cuando la consulta es pesada, y que normalmente
    se resuelven solos si se reintenta unos segundos después."""
    if isinstance(error, requests.exceptions.HTTPError) and error.response is not None:
        return error.response.status_code in (502, 503, 504)
    return isinstance(error, (requests.exceptions.ConnectionError, requests.exceptions.Timeout))


def llamar_con_reintentos(funcion, *args, **kwargs):
    """Ejecuta funcion(*args, **kwargs) reintentando con espera creciente (backoff
    exponencial) si el error es transitorio. Si el error no es transitorio, o si ya
    se agotaron los reintentos, deja que el error se propague normalmente."""
    ultimo_error = None
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            return funcion(*args, **kwargs)
        except Exception as error:
            ultimo_error = error
            if not _es_error_transitorio(error) or intento == MAX_REINTENTOS:
                raise
            espera = ESPERA_BASE_SEGUNDOS * (2 ** (intento - 1))
            print(f"  Aviso: el servidor de Geotab respondió con un error temporal ({error}). "
                  f"Reintentando en {espera}s (intento {intento}/{MAX_REINTENTOS})...")
            time.sleep(espera)
    raise ultimo_error


def obtener_zonas(client):
    zonas_raw = llamar_con_reintentos(client.get, 'Zone')
    zonas = []
    for z in zonas_raw:
        if isinstance(z, dict) and z.get('points'):
            poligono = [(p['x'], p['y']) for p in z['points']]
            lons = [p[0] for p in poligono]
            lats = [p[1] for p in poligono]
            area = 0
            n = len(poligono)
            for i in range(n):
                x1, y1 = poligono[i]
                x2, y2 = poligono[(i + 1) % n]
                area += x1 * y2 - x2 * y1
            area = abs(area) / 2
            zonas.append({'nombre': z.get('name', 'Zona sin nombre'), 'poligono': poligono, 'area': area})
    return zonas


def punto_en_poligono(lon, lat, poligono):
    n = len(poligono)
    dentro = False
    x1, y1 = poligono[0]
    for i in range(1, n + 1):
        x2, y2 = poligono[i % n]
        if lat > min(y1, y2) and lat <= max(y1, y2) and lon <= max(x1, x2):
            if y1 != y2:
                x_interseccion = (lat - y1) * (x2 - x1) / (y2 - y1) + x1
            if x1 == x2 or lon <= x_interseccion:
                dentro = not dentro
        x1, y1 = x2, y2
    return dentro


def determinar_zona(lon, lat, zonas):
    """Devuelve la zona MÁS ESPECÍFICA (de menor área) entre todas las que contengan el
    punto — igual que el tablero principal — para no quedarse con un barrio grande
    cuando en realidad cae dentro de una zona más puntual (ej. el predio del proveedor)."""
    if lon is None or lat is None:
        return 'Sin ubicación'
    candidatas = [z for z in zonas if punto_en_poligono(lon, lat, z['poligono'])]
    if not candidatas:
        return 'Fuera de zonas definidas'
    zona_mas_especifica = min(candidatas, key=lambda z: z['area'])
    return zona_mas_especifica['nombre']


# --- Parámetros del análisis (ajusta a tu gusto) ---
UMBRAL_CAIDA_PCT = 25        # puntos porcentuales de caída para considerarla sospechosa. Se subió
                              # de 15 a 25 porque el "piso de ruido" real del sensor de esta flota
                              # resultó ser ~30% (ver Categoria_Evento más abajo, que es el filtro
                              # principal — el umbral solo evita procesar caídas triviales de más).
UMBRAL_FALLA_SENSOR_PCT = 80  # caídas iguales o mayores a esto (ej. tanque casi lleno a casi vacío
                              # en minutos) no son físicamente creíbles como robo ni como consumo:
                              # se marcan aparte como probable falla/glitch del sensor.
VENTANA_MAXIMA_MINUTOS = 40  # la caída debe ocurrir en máximo estos minutos
DIAS_A_ANALIZAR = 15        # cuántos días hacia atrás analizar

ID_DIAGNOSTICO_COMBUSTIBLE = 'DiagnosticFuelLevelId'
ID_DIAGNOSTICO_ODOMETRO = 'DiagnosticOdometerId'  # Geotab lo entrega en metros
ID_DIAGNOSTICO_IGNICION = 'DiagnosticIgnitionId'  # 1 = encendido, 0 = apagado; Geotab solo
                                                   # registra un dato cuando el estado CAMBIA,
                                                   # así que cada fila es una transición real
ID_DIAGNOSTICO_ALTITUD = 'aZ_PCPTFQJUWGgwTodd5nhA'  # "Altitude Above Mean Sea Level". OPCIONAL:
                                                     # requiere un dispositivo Geotab GO9 o más
                                                     # reciente Y que Geotab/tu proveedor (Copiloto)
                                                     # lo habilite manualmente para tu cuenta — no
                                                     # viene activado por defecto. Si no está
                                                     # disponible, el script sigue funcionando sin
                                                     # la columna de pendiente aproximada.
DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR = 1.0  # por debajo de esto, se considera "prácticamente detenido"
TAMANO_LOTE_UBICACION = 50  # cuántas consultas de ubicación se agrupan en cada llamada a la API

REFERENCIA_MOTOR_POR_MARCA = {
    "volkswagen": "ISF 3.8",
    "volskwagen": "ISF 3.8",
    "mercedes": "OM926",
    "international": "L9",
    "foton": "X12",
    "kenworth": "ISM 11",
}

CIUDAD_A_FILTRAR = 'Bogotá'  # Cambia esto (o pon None) si quieres analizar otra ciudad o todas


def es_grupo_marca(nombre):
    nombre_l = nombre.strip().lower()
    return any(marca in nombre_l for marca in REFERENCIA_MOTOR_POR_MARCA)


def normalizar_ciudad(nombre):
    n = nombre.upper()
    if 'BOGOTA' in n or 'BOGOTÁ' in n:
        return 'Bogotá'
    if 'CALI' in n:
        return 'Cali'
    if 'VALLE' in n:
        return 'Valle'
    return nombre.strip()


def obtener_mapa_grupos(client):
    """Recorre la jerarquía de grupos de Geotab para saber a qué ciudad pertenece cada uno,
    exactamente igual a como lo hace el tablero de Streamlit."""
    todos_grupos = llamar_con_reintentos(client.get, 'Group')
    grupos_por_id = {g.get('id'): g for g in todos_grupos if isinstance(g, dict)}
    raiz = next((g for g in todos_grupos if g.get('name', '').strip().startswith('*')), None)
    if not raiz:
        return {}

    mapa = {}

    def obtener_id(referencia):
        return referencia['id'] if isinstance(referencia, dict) else referencia

    def recorrer(grupo_id, ciudad_actual):
        grupo_completo = grupos_por_id.get(grupo_id)
        if not grupo_completo:
            return
        mapa[grupo_id] = {'nombre': grupo_completo.get('name', ''), 'ciudad': ciudad_actual}
        for hijo in (grupo_completo.get('children') or []):
            recorrer(obtener_id(hijo), ciudad_actual)

    for hijo_raiz in (raiz.get('children') or []):
        hijo_id = obtener_id(hijo_raiz)
        hijo_completo = grupos_por_id.get(hijo_id, {})
        nombre = hijo_completo.get('name', '').strip()
        if es_grupo_marca(nombre):
            recorrer(hijo_id, 'Sin ciudad asignada')
        else:
            ciudad = normalizar_ciudad(nombre)
            recorrer(hijo_id, ciudad)
    return mapa


def resolver_ciudad(vehiculo, mapa_grupos):
    for g in (vehiculo.get('groups') or []):
        gid = g['id'] if isinstance(g, dict) else g
        info = mapa_grupos.get(gid)
        if info and info['ciudad'] and info['ciudad'] != 'Sin ciudad asignada':
            return info['ciudad']
    return 'Sin ciudad asignada'


def conectar():
    usuario = os.environ["GEOTAB_USUARIO"]
    contrasena = os.environ["GEOTAB_CONTRASENA"]
    base_de_datos = os.environ["GEOTAB_DATABASE"]
    servidor = os.environ.get("GEOTAB_SERVER", "my.geotab.com")
    client = mygeotab.API(username=usuario, password=contrasena,
                           database=base_de_datos, server=servidor)
    client.authenticate()
    return client


def obtener_vehiculos(client):
    dispositivos = llamar_con_reintentos(client.get, 'Device')
    mapa_grupos = obtener_mapa_grupos(client)

    resultado = {}
    for d in dispositivos:
        ciudad = resolver_ciudad(d, mapa_grupos)
        if CIUDAD_A_FILTRAR is not None and ciudad != CIUDAD_A_FILTRAR:
            continue
        resultado[d['id']] = d.get('name', d['id'])

    print(f"  {len(resultado)} de {len(dispositivos)} vehículos pertenecen a '{CIUDAD_A_FILTRAR}'.")
    return resultado


def extraer_niveles_combustible(client, dispositivos, f_inicio, f_fin):
    """Trae el nivel de combustible (%) de todos los vehículos en el rango dado."""
    calls = []
    for device_id in dispositivos:
        calls.append(('Get', {
            'typeName': 'StatusData',
            'search': {
                'deviceSearch': {'id': device_id},
                'diagnosticSearch': {'id': ID_DIAGNOSTICO_COMBUSTIBLE},
                'fromDate': f_inicio.isoformat(),
                'toDate': f_fin.isoformat(),
            }
        }))

    resultados = llamar_con_reintentos(client.multi_call, calls)

    filas = []
    for device_id, resultado in zip(dispositivos.keys(), resultados):
        for registro in resultado:
            filas.append({
                'Vehiculo': dispositivos[device_id],
                'DeviceId': device_id,
                'Fecha_Hora': pd.to_datetime(registro['dateTime']),
                'Nivel_Combustible': registro['data'],
            })

    if not filas:
        return pd.DataFrame(columns=['Vehiculo', 'DeviceId', 'Fecha_Hora', 'Nivel_Combustible'])

    return pd.DataFrame(filas)


def extraer_odometro(client, dispositivos, f_inicio, f_fin):
    """Trae el odómetro (metros) de todos los vehículos en el rango dado, para poder
    calcular cuánto se movió cada uno entre dos lecturas de combustible."""
    calls = []
    for device_id in dispositivos:
        calls.append(('Get', {
            'typeName': 'StatusData',
            'search': {
                'deviceSearch': {'id': device_id},
                'diagnosticSearch': {'id': ID_DIAGNOSTICO_ODOMETRO},
                'fromDate': f_inicio.isoformat(),
                'toDate': f_fin.isoformat(),
            }
        }))

    resultados = llamar_con_reintentos(client.multi_call, calls)

    filas = []
    for device_id, resultado in zip(dispositivos.keys(), resultados):
        for registro in resultado:
            filas.append({
                'Vehiculo': dispositivos[device_id],
                'Fecha_Hora': pd.to_datetime(registro['dateTime']),
                'Odometro_km': registro['data'] / 1000,
            })

    if not filas:
        return pd.DataFrame(columns=['Vehiculo', 'Fecha_Hora', 'Odometro_km'])

    return pd.DataFrame(filas).sort_values(['Vehiculo', 'Fecha_Hora'])


def extraer_ignicion(client, dispositivos, f_inicio, f_fin):
    """Trae el historial de encendido/apagado de todos los vehículos en el rango dado.
    Geotab solo genera un dato cuando el estado CAMBIA (no es una lectura continua),
    así que cada fila representa una transición real: se encendió o se apagó."""
    calls = []
    for device_id in dispositivos:
        calls.append(('Get', {
            'typeName': 'StatusData',
            'search': {
                'deviceSearch': {'id': device_id},
                'diagnosticSearch': {'id': ID_DIAGNOSTICO_IGNICION},
                'fromDate': f_inicio.isoformat(),
                'toDate': f_fin.isoformat(),
            }
        }))

    resultados = llamar_con_reintentos(client.multi_call, calls)

    filas = []
    for device_id, resultado in zip(dispositivos.keys(), resultados):
        for registro in resultado:
            filas.append({
                'Vehiculo': dispositivos[device_id],
                'Fecha_Hora': pd.to_datetime(registro['dateTime']),
                'Encendido': bool(registro['data']),
            })

    if not filas:
        return pd.DataFrame(columns=['Vehiculo', 'Fecha_Hora', 'Encendido'])

    return pd.DataFrame(filas).sort_values(['Vehiculo', 'Fecha_Hora'])


def extraer_altitud(client, dispositivos, f_inicio, f_fin):
    """Trae la altitud (metros sobre el nivel del mar) de cada vehículo, para poder
    aproximar la pendiente del terreno en cada evento — útil en flotas que operan en
    zonas montañosas, donde el combustible se mueve dentro del tanque (sloshing) en
    subidas/bajadas fuertes y puede generar caídas falsas.

    Este diagnóstico es OPCIONAL en Geotab: requiere un dispositivo GO9 o más reciente,
    y debe ser activado manualmente por Geotab o tu proveedor (Copiloto) — no viene
    activado por defecto en la mayoría de cuentas. Si no está disponible, esta función
    devuelve un DataFrame vacío en vez de fallar, y el resto del análisis sigue
    funcionando normalmente, solo que sin la columna de pendiente aproximada."""
    calls = []
    for device_id in dispositivos:
        calls.append(('Get', {
            'typeName': 'StatusData',
            'search': {
                'deviceSearch': {'id': device_id},
                'diagnosticSearch': {'id': ID_DIAGNOSTICO_ALTITUD},
                'fromDate': f_inicio.isoformat(),
                'toDate': f_fin.isoformat(),
            }
        }))

    try:
        resultados = llamar_con_reintentos(client.multi_call, calls)
    except Exception as error:
        print(f"  Aviso: no se pudo obtener altitud (posiblemente el diagnóstico no está "
              f"habilitado para esta cuenta/dispositivo): {error}")
        return pd.DataFrame(columns=['Vehiculo', 'Fecha_Hora', 'Altitud_m'])

    filas = []
    for device_id, resultado in zip(dispositivos.keys(), resultados):
        for registro in resultado:
            filas.append({
                'Vehiculo': dispositivos[device_id],
                'Fecha_Hora': pd.to_datetime(registro['dateTime']),
                'Altitud_m': registro['data'],
            })

    if not filas:
        print("  No se encontraron datos de altitud — probablemente el diagnóstico no está "
              "habilitado para esta cuenta o estos dispositivos. Pregúntale a Geotab/Copiloto "
              "si pueden activar 'Altitude Above Mean Sea Level' para tus vehículos.")
        return pd.DataFrame(columns=['Vehiculo', 'Fecha_Hora', 'Altitud_m'])

    return pd.DataFrame(filas).sort_values(['Vehiculo', 'Fecha_Hora'])


def agregar_distancia_recorrida(df_sospechosos, df_odometro):
    """Para cada evento sospechoso, calcula cuántos km recorrió el vehículo entre la
    lectura anterior y la lectura donde se detectó la caída, usando el odómetro más
    cercano disponible en cada momento (asof merge)."""
    if df_sospechosos.empty or df_odometro.empty:
        df_sospechosos = df_sospechosos.copy()
        df_sospechosos['Distancia_Recorrida_km'] = None
        df_sospechosos['Vehiculo_Practicamente_Detenido'] = None
        return df_sospechosos

    df = df_sospechosos.sort_values('Fecha_Anterior').copy()
    odo = df_odometro.sort_values('Fecha_Hora')

    antes = pd.merge_asof(
        df[['Vehiculo', 'Fecha_Anterior']].sort_values('Fecha_Anterior'),
        odo, left_on='Fecha_Anterior', right_on='Fecha_Hora',
        by='Vehiculo', direction='backward'
    )['Odometro_km']

    df_ord_despues = df.sort_values('Fecha_Hora')
    despues = pd.merge_asof(
        df_ord_despues[['Vehiculo', 'Fecha_Hora']],
        odo, on='Fecha_Hora', by='Vehiculo', direction='backward'
    )['Odometro_km']

    df = df.sort_values('Fecha_Anterior')
    df['Odometro_Antes_km'] = antes.values
    df = df.sort_values('Fecha_Hora')
    df['Odometro_Despues_km'] = despues.values

    df['Distancia_Recorrida_km'] = (df['Odometro_Despues_km'] - df['Odometro_Antes_km']).round(2)
    df['Vehiculo_Practicamente_Detenido'] = df['Distancia_Recorrida_km'] < DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR

    return df.drop(columns=['Odometro_Antes_km', 'Odometro_Despues_km'])


def agregar_pendiente_aproximada(df_sospechosos, df_altitud):
    """Aproxima la pendiente del terreno (% de desnivel) entre el inicio y el fin de cada
    evento: (cambio de altitud) / (distancia recorrida). Sirve para saber si la caída de
    combustible coincide con una subida o bajada fuerte — típico de zonas montañosas — que
    puede mover el combustible dentro del tanque (sloshing) y producir falsas alarmas.

    Si no hay datos de altitud disponibles (diagnóstico no habilitado, ver extraer_altitud),
    la columna queda vacía y el resto del análisis no se ve afectado."""
    if df_sospechosos.empty:
        return df_sospechosos

    df = df_sospechosos.copy()

    if df_altitud.empty or 'Distancia_Recorrida_km' not in df.columns:
        df['Pendiente_Aprox_Pct'] = None
        return df

    alt = df_altitud.sort_values('Fecha_Hora')

    antes = pd.merge_asof(
        df[['Vehiculo', 'Fecha_Anterior']].sort_values('Fecha_Anterior'),
        alt, left_on='Fecha_Anterior', right_on='Fecha_Hora',
        by='Vehiculo', direction='backward'
    )['Altitud_m']

    df_ord_despues = df.sort_values('Fecha_Hora')
    despues = pd.merge_asof(
        df_ord_despues[['Vehiculo', 'Fecha_Hora']],
        alt, on='Fecha_Hora', by='Vehiculo', direction='backward'
    )['Altitud_m']

    df = df.sort_values('Fecha_Anterior')
    df['Altitud_Antes_m'] = antes.values
    df = df.sort_values('Fecha_Hora')
    df['Altitud_Despues_m'] = despues.values

    distancia_m = df['Distancia_Recorrida_km'] * 1000
    desnivel_m = df['Altitud_Despues_m'] - df['Altitud_Antes_m']
    df['Pendiente_Aprox_Pct'] = None
    valido = distancia_m > 50  # ignora tramos muy cortos (<50 m), donde el % de pendiente es ruido de GPS
    df.loc[valido, 'Pendiente_Aprox_Pct'] = (desnivel_m[valido] / distancia_m[valido] * 100).round(1)

    return df.drop(columns=['Altitud_Antes_m', 'Altitud_Despues_m'])


def agregar_estado_ignicion(df_sospechosos, df_ignicion):
    """Para cada evento sospechoso, determina si el vehículo estaba encendido o
    apagado en ese momento, y a qué hora fue el último encendido y el último apagado
    antes del evento (usando las transiciones de ignición más cercanas, igual que
    agregar_distancia_recorrida hace con el odómetro)."""
    if df_sospechosos.empty:
        return df_sospechosos

    df = df_sospechosos.copy()

    if df_ignicion.empty:
        df['Vehiculo_Apagado'] = None
        df['Hora_Ultimo_Encendido'] = None
        df['Hora_Ultimo_Apagado'] = None
        return df

    encendidos = df_ignicion[df_ignicion['Encendido']][['Vehiculo', 'Fecha_Hora']].rename(
        columns={'Fecha_Hora': 'Hora_Ultimo_Encendido'}).sort_values('Hora_Ultimo_Encendido')
    apagados = df_ignicion[~df_ignicion['Encendido']][['Vehiculo', 'Fecha_Hora']].rename(
        columns={'Fecha_Hora': 'Hora_Ultimo_Apagado'}).sort_values('Hora_Ultimo_Apagado')

    df_ord = df.sort_values('Fecha_Hora').copy()

    df_ord['Hora_Ultimo_Encendido'] = pd.merge_asof(
        df_ord[['Vehiculo', 'Fecha_Hora']], encendidos,
        left_on='Fecha_Hora', right_on='Hora_Ultimo_Encendido',
        by='Vehiculo', direction='backward'
    )['Hora_Ultimo_Encendido'].values

    df_ord['Hora_Ultimo_Apagado'] = pd.merge_asof(
        df_ord[['Vehiculo', 'Fecha_Hora']], apagados,
        left_on='Fecha_Hora', right_on='Hora_Ultimo_Apagado',
        by='Vehiculo', direction='backward'
    )['Hora_Ultimo_Apagado'].values

    def _estaba_apagado(fila):
        hay_encendido = pd.notna(fila['Hora_Ultimo_Encendido'])
        hay_apagado = pd.notna(fila['Hora_Ultimo_Apagado'])
        if not hay_encendido and not hay_apagado:
            return None  # sin datos de ignición para este vehículo en el periodo
        if not hay_apagado:
            return False
        if not hay_encendido:
            return True
        # el último evento (el más reciente de los dos) define el estado actual
        return fila['Hora_Ultimo_Apagado'] > fila['Hora_Ultimo_Encendido']

    df_ord['Vehiculo_Apagado'] = df_ord.apply(_estaba_apagado, axis=1)

    df = df.sort_values('Fecha_Hora')
    df['Vehiculo_Apagado'] = df_ord['Vehiculo_Apagado'].values
    df['Hora_Ultimo_Encendido'] = df_ord['Hora_Ultimo_Encendido'].values
    df['Hora_Ultimo_Apagado'] = df_ord['Hora_Ultimo_Apagado'].values

    return df.sort_values('Caida_Pct', ascending=False)


def clasificar_eventos(df_sospechosos, umbral_falla_sensor=UMBRAL_FALLA_SENSOR_PCT):
    """Clasifica cada evento en una categoría, en vez de dejar que el usuario adivine con
    solo la magnitud de la caída:

    - 'Revisar sensor (posible falla)': caída extrema (>= umbral_falla_sensor). Vaciar el
      tanque casi por completo en minutos no es creíble ni por consumo normal ni por robo
      — apunta a una falla/glitch del sensor de ese vehículo puntual.
    - 'Robo probable': el vehículo estaba detenido y/o con el motor apagado. El consumo
      normal de motor no puede explicar una caída así — es la señal más creíble de robo.
    - 'Posible ruido de sensor / pendiente': el vehículo estaba en movimiento normal. Puede
      ser una caída real, pero también es el patrón típico de sloshing del combustible en
      pendientes, frenadas o terreno irregular — amerita revisión pero con menor prioridad."""
    if df_sospechosos.empty:
        return df_sospechosos

    df = df_sospechosos.copy()

    def _clasificar(fila):
        if fila['Caida_Pct'] >= umbral_falla_sensor:
            return 'Revisar sensor (posible falla)'
        if fila.get('Vehiculo_Practicamente_Detenido') is True or fila.get('Vehiculo_Apagado') is True:
            return 'Robo probable'
        return 'Posible ruido de sensor / pendiente'

    df['Categoria_Evento'] = df.apply(_clasificar, axis=1)
    return df


def detectar_caidas_sospechosas(df_combustible, umbral_pct=UMBRAL_CAIDA_PCT,
                                 ventana_min=VENTANA_MAXIMA_MINUTOS):
    """Marca como sospechosa cualquier caída >= umbral_pct dentro de una ventana de
    ventana_min minutos, usando el nivel MÁS BAJO alcanzado en ese lapso (no solo la
    lectura inmediatamente anterior). Esto captura tanto caídas de un solo golpe como
    caídas repartidas en varios pasos pequeños dentro de la misma ventana."""
    if df_combustible.empty:
        return pd.DataFrame()

    eventos = []
    for vehiculo, grupo in df_combustible.groupby('Vehiculo'):
        grupo = grupo.sort_values('Fecha_Hora').reset_index(drop=True)
        if len(grupo) < 2:
            continue

        device_id = grupo['DeviceId'].iloc[0] if 'DeviceId' in grupo.columns else None

        # Truco para calcular un "mínimo hacia adelante en el tiempo" con pandas:
        # invertimos el eje de tiempo (reflejado sobre un pivote) y usamos un rolling
        # normal (hacia atrás) sobre ese eje invertido, que equivale a mirar hacia
        # adelante en el tiempo real.
        pivote = grupo['Fecha_Hora'].max() + pd.Timedelta(minutes=1)
        grupo['rev_time'] = pivote - grupo['Fecha_Hora']
        grupo_rev = grupo.sort_values('rev_time').set_index('rev_time')
        min_forward = grupo_rev['Nivel_Combustible'].rolling(f'{ventana_min}min').min()
        grupo_rev['Nivel_Minimo_En_Ventana'] = min_forward
        grupo = grupo_rev.reset_index(drop=True)

        grupo['Caida_Pct'] = grupo['Nivel_Combustible'] - grupo['Nivel_Minimo_En_Ventana']
        grupo['Sospechoso'] = grupo['Caida_Pct'] >= umbral_pct

        # Agrupa rachas consecutivas marcadas como sospechosas (parte de la misma caída)
        # y se queda con el punto de mayor caída de cada racha, para no duplicar el
        # mismo evento varias veces.
        grupo['grupo_evento'] = (grupo['Sospechoso'] != grupo['Sospechoso'].shift()).cumsum()
        for _, sub in grupo[grupo['Sospechoso']].groupby('grupo_evento'):
            fila_peor = sub.loc[sub['Caida_Pct'].idxmax()]

            # Busca, dentro de la ventana real (no un desplazamiento fijo), la lectura
            # exacta donde se tocó el nivel mínimo — así Fecha_Hora queda en el momento
            # real de la caída y no en un punto artificial (inicio + ventana_min), que es
            # lo que usan después agregar_ubicacion_a_eventos y agregar_estado_ignicion
            # para saber dónde estaba y si estaba encendido el vehículo.
            ventana_futuro = grupo[
                (grupo['Fecha_Hora'] >= fila_peor['Fecha_Hora']) &
                (grupo['Fecha_Hora'] <= fila_peor['Fecha_Hora'] + pd.Timedelta(minutes=ventana_min))
            ]
            fila_minima = ventana_futuro.loc[ventana_futuro['Nivel_Combustible'].idxmin()]
            momento_minimo = fila_minima['Fecha_Hora']

            eventos.append({
                'Vehiculo': vehiculo,
                'DeviceId': device_id,
                'Fecha_Anterior': fila_peor['Fecha_Hora'],
                'Fecha_Hora': momento_minimo,
                'Nivel_Anterior': fila_peor['Nivel_Combustible'],
                'Nivel_Combustible': fila_peor['Nivel_Minimo_En_Ventana'],
                'Caida_Pct': fila_peor['Caida_Pct'],
                'Minutos_Transcurridos': round(
                    (momento_minimo - fila_peor['Fecha_Hora']).total_seconds() / 60, 1),
            })

    if not eventos:
        return pd.DataFrame()

    return pd.DataFrame(eventos)[[
        'Vehiculo', 'DeviceId', 'Fecha_Anterior', 'Fecha_Hora', 'Nivel_Anterior',
        'Nivel_Combustible', 'Caida_Pct', 'Minutos_Transcurridos'
    ]].sort_values('Caida_Pct', ascending=False)


def agregar_ubicacion_a_eventos(client, df_sospechosos, zonas):
    """Para cada evento sospechoso, busca la posición GPS real del vehículo en ese
    momento y determina en qué zona de Geotab cae (ej. el taller/proveedor).

    Agrupa las consultas en lotes de TAMANO_LOTE_UBICACION usando multi_call, en vez
    de hacer una llamada HTTP separada por cada evento: con miles de eventos, una
    llamada por evento satura el servidor del proveedor y termina en timeouts (504)."""
    if df_sospechosos.empty:
        return df_sospechosos

    filas = list(df_sospechosos.itertuples())
    zonas_detectadas = [None] * len(filas)
    cerca_proveedor = [None] * len(filas)
    latitudes = [None] * len(filas)
    longitudes = [None] * len(filas)
    total = len(filas)

    for inicio in range(0, total, TAMANO_LOTE_UBICACION):
        lote = filas[inicio: inicio + TAMANO_LOTE_UBICACION]
        calls = []
        for fila in lote:
            desde = (fila.Fecha_Hora - pd.Timedelta(minutes=2)).isoformat()
            hasta = (fila.Fecha_Hora + pd.Timedelta(minutes=2)).isoformat()
            calls.append(('Get', {
                'typeName': 'LogRecord',
                'search': {
                    'deviceSearch': {'id': fila.DeviceId},
                    'fromDate': desde,
                    'toDate': hasta,
                }
            }))

        resultados = llamar_con_reintentos(client.multi_call, calls)

        for offset, (fila, registros) in enumerate(zip(lote, resultados)):
            idx = inicio + offset
            if registros:
                mas_cercano = min(registros, key=lambda r: abs(pd.to_datetime(r['dateTime']) - fila.Fecha_Hora))
                lon, lat = mas_cercano.get('longitude'), mas_cercano.get('latitude')
            else:
                lon, lat = None, None
            zona = determinar_zona(lon, lat, zonas)
            zonas_detectadas[idx] = zona
            cerca_proveedor[idx] = any(z.upper() in zona.upper() for z in ZONAS_SOSPECHOSAS)
            latitudes[idx] = lat
            longitudes[idx] = lon

        print(f"  Validando ubicación: {min(inicio + TAMANO_LOTE_UBICACION, total)}/{total}...")

    df_sospechosos = df_sospechosos.copy()
    df_sospechosos['Zona_Detectada'] = zonas_detectadas
    df_sospechosos['Cerca_Proveedor_Sospechoso'] = cerca_proveedor
    df_sospechosos['Latitud'] = latitudes
    df_sospechosos['Longitud'] = longitudes
    df_sospechosos['Enlace_Maps'] = [
        f"https://www.google.com/maps?q={lat},{lon}" if pd.notna(lat) and pd.notna(lon) else None
        for lat, lon in zip(latitudes, longitudes)
    ]
    return df_sospechosos


# Nominatim (OpenStreetMap) es un servicio GRATUITO de geocodificación inversa, pero de uso
# limitado (máx. ~1 consulta/segundo) y exige identificarte con un user-agent válido. Por eso
# solo se usa para los eventos prioritarios (Robo probable / Revisar sensor), no para los miles
# de eventos de ruido. Si tu volumen de casos prioritarios crece mucho, o quieres direcciones
# para TODOS los eventos, lo ideal es migrar a la API de Geocodificación de Google (de pago).
NOMINATIM_USER_AGENT = "AnalisisRoboCombustiblePromoambiental/1.0 (cambia-este-correo@promoambientaldistrito.com)"
NOMINATIM_ESPERA_SEGUNDOS = 1.1  # respeta el límite de uso justo (fair use) de Nominatim


def obtener_direccion_desde_coordenadas(lat, lon):
    """Convierte una coordenada en una dirección de texto legible (geocodificación inversa)
    usando Nominatim/OpenStreetMap. Si falla (sin internet, límite de uso, etc.) devuelve
    None en vez de detener el script — la dirección es un dato adicional, no crítico."""
    try:
        respuesta = requests.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={'lat': lat, 'lon': lon, 'format': 'jsonv2'},
            headers={'User-Agent': NOMINATIM_USER_AGENT},
            timeout=10,
        )
        respuesta.raise_for_status()
        return respuesta.json().get('display_name')
    except Exception as error:
        print(f"    Aviso: no se pudo obtener la dirección para ({lat}, {lon}): {error}")
        return None


def agregar_direccion_aproximada(df_sospechosos):
    """Agrega una dirección de texto legible ('Direccion_Aproximada') a los eventos más
    prioritarios (Robo probable y Revisar sensor), consultando Nominatim una fila a la vez
    con una pausa entre consultas para respetar su límite de uso."""
    if df_sospechosos.empty or 'Categoria_Evento' not in df_sospechosos.columns:
        return df_sospechosos

    df = df_sospechosos.copy()
    df['Direccion_Aproximada'] = None

    prioritarios = df[df['Categoria_Evento'].isin(['Robo probable', 'Revisar sensor (posible falla)'])]
    prioritarios = prioritarios[prioritarios['Latitud'].notna() & prioritarios['Longitud'].notna()]

    total = len(prioritarios)
    if total == 0:
        return df

    print(f"  Buscando dirección de texto para los {total} eventos prioritarios (puede tardar ~1s por evento)...")
    for i, (idx, fila) in enumerate(prioritarios.iterrows(), start=1):
        direccion = obtener_direccion_desde_coordenadas(fila['Latitud'], fila['Longitud'])
        df.at[idx, 'Direccion_Aproximada'] = direccion
        if i % 10 == 0 or i == total:
            print(f"    Direcciones: {i}/{total}...")
        if i < total:
            time.sleep(NOMINATIM_ESPERA_SEGUNDOS)

    return df


def resumen_por_vehiculo(df_sospechosos):
    """Cuenta cuántos eventos sospechosos tiene cada vehículo, para ver si se repite."""
    if df_sospechosos.empty:
        return pd.DataFrame()

    agregaciones = {
        'Eventos_Sospechosos': ('Caida_Pct', 'count'),
        'Caida_Promedio_Pct': ('Caida_Pct', 'mean'),
        'Caida_Maxima_Pct': ('Caida_Pct', 'max'),
        'Primer_Evento': ('Fecha_Hora', 'min'),
        'Ultimo_Evento': ('Fecha_Hora', 'max'),
    }
    if 'Cerca_Proveedor_Sospechoso' in df_sospechosos.columns:
        agregaciones['Eventos_Cerca_Proveedor'] = ('Cerca_Proveedor_Sospechoso', 'sum')
    if 'Vehiculo_Practicamente_Detenido' in df_sospechosos.columns:
        agregaciones['Eventos_Vehiculo_Detenido'] = ('Vehiculo_Practicamente_Detenido', 'sum')
    if 'Vehiculo_Apagado' in df_sospechosos.columns:
        agregaciones['Eventos_Vehiculo_Apagado'] = ('Vehiculo_Apagado', 'sum')
    if 'Categoria_Evento' in df_sospechosos.columns:
        agregaciones['Eventos_Robo_Probable'] = (
            'Categoria_Evento', lambda s: (s == 'Robo probable').sum())
        agregaciones['Eventos_Revisar_Sensor'] = (
            'Categoria_Evento', lambda s: (s == 'Revisar sensor (posible falla)').sum())

    resumen = df_sospechosos.groupby('Vehiculo').agg(**agregaciones).sort_values(
        'Eventos_Sospechosos', ascending=False
    )

    return resumen.reset_index()


def calcular_datos_panel(detalle):
    """Calcula los totales y series que alimentan las gráficas del Panel General:
    % de eventos cerca del proveedor, % con motor apagado, % con vehículo detenido,
    y cuántos eventos sospechosos hubo por día (para ver si la tendencia sube o baja)."""
    if detalle.empty:
        return None

    total = len(detalle)

    def contar_si(col):
        if col not in detalle.columns:
            return None
        return int((detalle[col] == 'Sí').sum())

    eventos_por_dia = (
        detalle.assign(Dia=pd.to_datetime(detalle['Fecha_Hora']).dt.date)
        .groupby('Dia').size().reset_index(name='Eventos')
        .sort_values('Dia')
    )
    eventos_por_dia['Dia'] = eventos_por_dia['Dia'].astype(str)

    por_categoria = None
    if 'Categoria_Evento' in detalle.columns:
        conteo = detalle['Categoria_Evento'].value_counts()
        orden = ['Robo probable', 'Revisar sensor (posible falla)', 'Posible ruido de sensor / pendiente']
        por_categoria = [(cat, int(conteo.get(cat, 0))) for cat in orden if conteo.get(cat, 0) > 0]

    return {
        'total': total,
        'cerca_proveedor': contar_si('Cerca_Proveedor_Sospechoso'),
        'apagado': contar_si('Vehiculo_Apagado'),
        'detenido': contar_si('Vehiculo_Practicamente_Detenido'),
        'eventos_por_dia': eventos_por_dia,
        'por_categoria': por_categoria,
    }


def generar_excel_formateado(nombre_archivo, resumen, detalle):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference

    COLOR_ENCABEZADO = "1F4E4A"   # verde oscuro, igual al del tablero
    COLOR_TEXTO_ENCABEZADO = "FFFFFF"
    COLOR_RESALTADO_PROVEEDOR = "FFF3B0"   # amarillo suave: cerca del proveedor sospechoso
    COLOR_RESALTADO_DETENIDO = "F4B4AE"    # rojo suave: robo probable (detenido y/o apagado)
    COLOR_RESALTADO_FALLA_SENSOR = "FFD966"  # ámbar: caída extrema, probable falla del sensor
    COLOR_BORDE = "D9D9D9"

    fuente_encabezado = Font(bold=True, color=COLOR_TEXTO_ENCABEZADO, size=11)
    relleno_encabezado = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
    relleno_proveedor = PatternFill("solid", fgColor=COLOR_RESALTADO_PROVEEDOR)
    relleno_detenido = PatternFill("solid", fgColor=COLOR_RESALTADO_DETENIDO)
    relleno_falla_sensor = PatternFill("solid", fgColor=COLOR_RESALTADO_FALLA_SENSOR)
    borde_delgado = Border(*(Side(style='thin', color=COLOR_BORDE) for _ in range(4)))
    alineacion_centrada = Alignment(horizontal='center', vertical='center')

    wb = Workbook()

    # --- Hoja 1: Léame, explicando el análisis en lenguaje sencillo ---
    ws_intro = wb.active
    ws_intro.title = "Léame"
    ws_intro.column_dimensions['A'].width = 100
    lineas_intro = [
        ("ANÁLISIS DE POSIBLE ROBO DE COMBUSTIBLE", True, 14),
        ("", False, 11),
        (f"Periodo analizado: {fecha_inicio.date()} al {fecha_fin.date()}", False, 11),
        (f"Ciudad: {CIUDAD_A_FILTRAR or 'Todas'}", False, 11),
        ("", False, 11),
        ("¿Qué hace este análisis?", True, 12),
        ("Revisa el nivel de combustible (%) de cada vehículo y marca como 'sospechoso'", False, 11),
        (f"cualquier caída de {UMBRAL_CAIDA_PCT} puntos porcentuales o más, ocurrida en", False, 11),
        (f"{VENTANA_MAXIMA_MINUTOS} minutos o menos.", False, 11),
        ("", False, 11),
        ("La hoja 'Panel General' resume todo de un vistazo: totales, y gráficas de", True, 11),
        ("categoría, estado del motor, ubicación de los eventos y su tendencia día a día.", False, 11),
        ("", False, 11),
        ("¿Cómo leer la hoja 'Resumen por vehiculo'?", True, 12),
        ("Cada fila es un vehículo. Entre más eventos sospechosos tenga, más veces se le", False, 11),
        ("detectó una caída brusca de combustible en el periodo analizado.", False, 11),
        ("", False, 11),
        ("¿Cómo leer la hoja 'Detalle de eventos'?", True, 12),
        ("Cada fila es una caída puntual detectada, con fecha, hora, cuánto bajó el", False, 11),
        ("combustible y en cuántos minutos. La columna 'Zona_Detectada' indica en qué", False, 11),
        ("zona se encontraba el vehículo, y 'Distancia_Recorrida_km' cuánto se movió", False, 11),
        ("entre esas dos lecturas.", False, 11),
        ("", False, 11),
        ("Las columnas 'Vehiculo_Apagado', 'Hora_Ultimo_Encendido' y", True, 11),
        ("'Hora_Ultimo_Apagado' indican si el motor estaba encendido o apagado en el", False, 11),
        ("momento de la caída, y a qué hora ocurrió el encendido/apagado más reciente", False, 11),
        ("antes de ese momento. Un vehículo APAGADO perdiendo combustible es la señal", False, 11),
        ("más fuerte de robo: el consumo normal del motor no puede explicar la caída.", False, 11),
        ("", False, 11),
        ("Ubicación exacta: la columna 'Enlace_Maps' es un enlace ('Ver en Maps') que", True, 11),
        ("abre directamente el punto GPS del evento en Google Maps — no hay que copiar", False, 11),
        ("ninguna dirección, solo hacer clic. Para los eventos más prioritarios (Robo", False, 11),
        ("probable y Revisar sensor) además se busca una 'Direccion_Aproximada' en texto,", False, 11),
        ("útil si necesitas compartirla por WhatsApp o pegarla en otro sistema.", False, 11),
        ("", False, 11),
        ("Si tu proveedor tiene habilitado el diagnóstico de altitud, verás además", True, 11),
        ("'Pendiente_Aprox_Pct': el desnivel aproximado del terreno durante el evento.", False, 11),
        ("Pendientes fuertes pueden mover el combustible dentro del tanque (sloshing)", False, 11),
        ("y generar caídas falsas — típico en flotas que operan en zona montañosa.", False, 11),
        ("", False, 11),
        ("¿Qué significa la columna 'Categoria_Evento'?", True, 12),
        ("Cada evento se clasifica según qué tan creíble es como robo real, no solo por", False, 11),
        ("el tamaño de la caída:", False, 11),
        ("🔴 'Robo probable': el vehículo estaba detenido y/o con el motor apagado — el", False, 11),
        ("consumo normal de motor no explica la caída. Revisa esto primero (hoja", False, 11),
        ("'Casos mas probables').", False, 11),
        ("🟠 'Revisar sensor (posible falla)': caída extrema "
         f"(≥ {UMBRAL_FALLA_SENSOR_PCT}%), como vaciar", False, 11),
        ("el tanque casi por completo en minutos. No es creíble ni como consumo ni", False, 11),
        ("como robo — apunta a una falla del sensor de ese vehículo puntual (hoja", False, 11),
        ("'Revisar sensor').", False, 11),
        ("⚪ 'Posible ruido de sensor / pendiente': el vehículo se estaba moviendo con", False, 11),
        ("normalidad. Puede ser real, pero también es el patrón típico de ruido del", False, 11),
        ("sensor por vibración o pendiente — menor prioridad.", False, 11),
        ("", False, 11),
        ("🟡 Amarillo (además de los colores anteriores): ocurrió cerca de una zona", False, 11),
        (f"sospechosa ({', '.join(ZONAS_SOSPECHOSAS)}).", False, 11),
        ("", False, 11),
        ("La hoja 'Eventos en zona sospechosa' es un filtro: solo trae las filas que", False, 11),
        ("ocurrieron dentro de esas zonas, para revisarlas sin tener que buscarlas.", False, 11),
        ("", False, 11),
        ("⚠️ Importante:", True, 12),
        ("Una caída detectada NO confirma un robo por sí sola — puede deberse también a", False, 11),
        ("ruido del sensor de combustible. Prioriza la hoja 'Casos mas probables',", False, 11),
        ("y valida contra el historial de viajes real antes de tomar acción.", False, 11),
    ]
    for i, (texto, negrita, tamano) in enumerate(lineas_intro, start=1):
        celda = ws_intro.cell(row=i, column=1, value=texto)
        celda.font = Font(bold=negrita, size=tamano,
                           color=COLOR_ENCABEZADO if negrita else "000000")
        celda.alignment = Alignment(wrap_text=True, vertical='center')

    # --- Hoja 2: Panel General, con KPIs y gráficas de resumen ---
    datos_panel = calcular_datos_panel(detalle)
    ws_panel = wb.create_sheet("Panel General")
    if datos_panel:
        ws_panel.column_dimensions['A'].width = 34
        ws_panel.column_dimensions['B'].width = 14

        ws_panel['A1'] = "PANEL GENERAL"
        ws_panel['A1'].font = Font(bold=True, size=14, color=COLOR_ENCABEZADO)

        kpis = [
            ("Total de eventos sospechosos", datos_panel['total']),
            ("Eventos cerca de zona sospechosa", datos_panel['cerca_proveedor']),
            ("Eventos con motor apagado", datos_panel['apagado']),
            ("Eventos con vehículo detenido", datos_panel['detenido']),
        ]
        fila = 3
        for etiqueta, valor in kpis:
            if valor is None:
                continue
            ws_panel.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
            ws_panel.cell(row=fila, column=2, value=valor)
            fila += 1

        # --- Tabla auxiliar para el gráfico circular "Categoría del evento" ---
        fila_categoria_inicio = fila + 2
        if datos_panel['por_categoria']:
            ws_panel.cell(row=fila_categoria_inicio, column=1, value="Categoría del evento").font = Font(bold=True)
            for i, (cat, cantidad) in enumerate(datos_panel['por_categoria'], start=1):
                ws_panel.cell(row=fila_categoria_inicio + i, column=1, value=cat)
                ws_panel.cell(row=fila_categoria_inicio + i, column=2, value=cantidad)
            fila_siguiente = fila_categoria_inicio + len(datos_panel['por_categoria']) + 2
        else:
            fila_siguiente = fila_categoria_inicio

        # --- Tabla auxiliar para el gráfico circular "Estado del motor" ---
        fila_estado_motor = fila_siguiente
        ws_panel.cell(row=fila_estado_motor, column=1, value="Estado del motor").font = Font(bold=True)
        ws_panel.cell(row=fila_estado_motor + 1, column=1, value="Apagado")
        ws_panel.cell(row=fila_estado_motor + 1, column=2, value=datos_panel['apagado'] or 0)
        ws_panel.cell(row=fila_estado_motor + 2, column=1, value="Encendido")
        ws_panel.cell(row=fila_estado_motor + 2, column=2,
                       value=datos_panel['total'] - (datos_panel['apagado'] or 0))

        # --- Tabla auxiliar para el gráfico circular "Ubicación del evento" ---
        fila_ubicacion = fila_estado_motor + 4
        ws_panel.cell(row=fila_ubicacion, column=1, value="Ubicación del evento").font = Font(bold=True)
        ws_panel.cell(row=fila_ubicacion + 1, column=1, value="Cerca de zona sospechosa")
        ws_panel.cell(row=fila_ubicacion + 1, column=2, value=datos_panel['cerca_proveedor'] or 0)
        ws_panel.cell(row=fila_ubicacion + 2, column=1, value="Fuera de zona sospechosa")
        ws_panel.cell(row=fila_ubicacion + 2, column=2,
                       value=datos_panel['total'] - (datos_panel['cerca_proveedor'] or 0))

        # --- Tabla auxiliar para el gráfico de tendencia "Eventos por día" ---
        fila_tendencia_inicio = fila_ubicacion + 4
        ws_panel.cell(row=fila_tendencia_inicio, column=1, value="Día").font = Font(bold=True)
        ws_panel.cell(row=fila_tendencia_inicio, column=2, value="Eventos").font = Font(bold=True)
        eventos_por_dia = datos_panel['eventos_por_dia']
        for i, (_, r) in enumerate(eventos_por_dia.iterrows(), start=1):
            ws_panel.cell(row=fila_tendencia_inicio + i, column=1, value=r['Dia'])
            ws_panel.cell(row=fila_tendencia_inicio + i, column=2, value=int(r['Eventos']))
        fila_tendencia_fin = fila_tendencia_inicio + len(eventos_por_dia)

        # Gráfico circular: categoría del evento (robo probable / revisar sensor / ruido) —
        # el más importante, así que va primero y más grande
        if datos_panel['por_categoria']:
            pie_categoria = PieChart()
            pie_categoria.title = "Categoría de los eventos"
            datos = Reference(ws_panel, min_col=2, min_row=fila_categoria_inicio + 1,
                               max_row=fila_categoria_inicio + len(datos_panel['por_categoria']))
            categorias = Reference(ws_panel, min_col=1, min_row=fila_categoria_inicio + 1,
                                    max_row=fila_categoria_inicio + len(datos_panel['por_categoria']))
            pie_categoria.add_data(datos)
            pie_categoria.set_categories(categorias)
            pie_categoria.height = 9
            pie_categoria.width = 14
            ws_panel.add_chart(pie_categoria, "D2")

        # Gráfico circular: estado del motor (encendido vs apagado)
        pie_motor = PieChart()
        pie_motor.title = "Estado del motor en los eventos"
        datos = Reference(ws_panel, min_col=2, min_row=fila_estado_motor + 1, max_row=fila_estado_motor + 2)
        categorias = Reference(ws_panel, min_col=1, min_row=fila_estado_motor + 1, max_row=fila_estado_motor + 2)
        pie_motor.add_data(datos)
        pie_motor.set_categories(categorias)
        pie_motor.height = 8
        pie_motor.width = 12
        ws_panel.add_chart(pie_motor, "L2")

        # Gráfico circular: ubicación (cerca o no de zona sospechosa)
        pie_ubicacion = PieChart()
        pie_ubicacion.title = "Ubicación de los eventos"
        datos = Reference(ws_panel, min_col=2, min_row=fila_ubicacion + 1, max_row=fila_ubicacion + 2)
        categorias = Reference(ws_panel, min_col=1, min_row=fila_ubicacion + 1, max_row=fila_ubicacion + 2)
        pie_ubicacion.add_data(datos)
        pie_ubicacion.set_categories(categorias)
        pie_ubicacion.height = 8
        pie_ubicacion.width = 12
        ws_panel.add_chart(pie_ubicacion, "S2")

        # Gráfico de barras: eventos sospechosos por día, para ver la tendencia en el tiempo
        if len(eventos_por_dia) > 0:
            barras_tendencia = BarChart()
            barras_tendencia.type = "col"
            barras_tendencia.title = "Eventos sospechosos por día"
            barras_tendencia.y_axis.title = "Eventos"
            barras_tendencia.x_axis.title = "Día"
            barras_tendencia.style = 10
            datos = Reference(ws_panel, min_col=2, min_row=fila_tendencia_inicio, max_row=fila_tendencia_fin)
            categorias = Reference(ws_panel, min_col=1, min_row=fila_tendencia_inicio + 1, max_row=fila_tendencia_fin)
            barras_tendencia.add_data(datos, titles_from_data=True)
            barras_tendencia.set_categories(categorias)
            barras_tendencia.width = 30
            barras_tendencia.height = 10
            ws_panel.add_chart(barras_tendencia, "D20")
    else:
        ws_panel['A1'] = "No se detectaron eventos sospechosos en el periodo analizado."

    # --- Hojas de datos ---
    hojas_a_generar = [("Resumen por vehiculo", resumen), ("Detalle de eventos", detalle)]
    if 'Categoria_Evento' in detalle.columns:
        df_robo_probable = detalle[detalle['Categoria_Evento'] == 'Robo probable'].copy()
        hojas_a_generar.append(("Casos mas probables", df_robo_probable))
        df_revisar_sensor = detalle[detalle['Categoria_Evento'] == 'Revisar sensor (posible falla)'].copy()
        if not df_revisar_sensor.empty:
            hojas_a_generar.append(("Revisar sensor", df_revisar_sensor))
    if 'Cerca_Proveedor_Sospechoso' in detalle.columns:
        df_filtro_zona = detalle[detalle['Cerca_Proveedor_Sospechoso'] == 'Sí'].copy()
        hojas_a_generar.append(("Eventos en zona sospechosa", df_filtro_zona))

    hojas_creadas = {}
    for nombre_hoja, df in hojas_a_generar:
        ws = wb.create_sheet(nombre_hoja)
        hojas_creadas[nombre_hoja] = ws
        ws.append(list(df.columns))
        for celda in ws[1]:
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado
            celda.alignment = alineacion_centrada
            celda.border = borde_delgado
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        col_categoria = None
        if 'Categoria_Evento' in df.columns:
            col_categoria = list(df.columns).index('Categoria_Evento') + 1
        col_proveedor = None
        if 'Cerca_Proveedor_Sospechoso' in df.columns:
            col_proveedor = list(df.columns).index('Cerca_Proveedor_Sospechoso') + 1
        col_maps = None
        if 'Enlace_Maps' in df.columns:
            col_maps = list(df.columns).index('Enlace_Maps') + 1

        for fila in df.itertuples(index=False):
            ws.append(list(fila))
            fila_actual = ws.max_row
            categoria = ws.cell(row=fila_actual, column=col_categoria).value if col_categoria else None
            es_proveedor = col_proveedor and ws.cell(row=fila_actual, column=col_proveedor).value == 'Sí'
            for celda in ws[fila_actual]:
                celda.border = borde_delgado
                if categoria == 'Robo probable':
                    celda.fill = relleno_detenido
                elif categoria == 'Revisar sensor (posible falla)':
                    celda.fill = relleno_falla_sensor
                elif es_proveedor:
                    celda.fill = relleno_proveedor

            if col_maps:
                celda_maps = ws.cell(row=fila_actual, column=col_maps)
                if celda_maps.value:
                    celda_maps.hyperlink = celda_maps.value
                    celda_maps.value = "Ver en Maps"
                    celda_maps.font = Font(color="1155CC", underline="single")

        # Ajustar ancho de columnas según el contenido
        for i, col in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            ancho_maximo = max(
                [len(str(col))] + [len(str(v)) for v in df[col].astype(str).values]
            )
            ws.column_dimensions[letra].width = min(ancho_maximo + 4, 45)

    # Gráfico de barras en "Resumen por vehiculo": los vehículos con más eventos sospechosos
    if 'Resumen por vehiculo' in hojas_creadas and not resumen.empty and \
            'Eventos_Sospechosos' in resumen.columns:
        ws_resumen = hojas_creadas['Resumen por vehiculo']
        col_vehiculo = list(resumen.columns).index('Vehiculo') + 1
        col_eventos = list(resumen.columns).index('Eventos_Sospechosos') + 1
        top_n = min(15, len(resumen))  # el resumen ya viene ordenado de más a menos eventos

        barras_vehiculos = BarChart()
        barras_vehiculos.type = "col"
        barras_vehiculos.title = "Vehículos con más eventos sospechosos"
        barras_vehiculos.y_axis.title = "Eventos sospechosos"
        barras_vehiculos.x_axis.title = "Vehículo"
        barras_vehiculos.style = 10
        datos = Reference(ws_resumen, min_col=col_eventos, min_row=1, max_row=1 + top_n)
        categorias = Reference(ws_resumen, min_col=col_vehiculo, min_row=2, max_row=1 + top_n)
        barras_vehiculos.add_data(datos, titles_from_data=True)
        barras_vehiculos.set_categories(categorias)
        barras_vehiculos.width = 26
        barras_vehiculos.height = 12
        ws_resumen.add_chart(barras_vehiculos, f"{get_column_letter(len(resumen.columns) + 2)}2")

    wb.save(nombre_archivo)


if __name__ == '__main__':
    print("Conectando a Geotab...")
    client = conectar()

    fecha_fin = datetime.datetime.now(datetime.UTC)
    fecha_inicio = fecha_fin - datetime.timedelta(days=DIAS_A_ANALIZAR)

    print(f"Trayendo niveles de combustible de los últimos {DIAS_A_ANALIZAR} días...")
    dispositivos = obtener_vehiculos(client)
    df_combustible = extraer_niveles_combustible(client, dispositivos, fecha_inicio, fecha_fin)
    print(f"  {len(df_combustible)} lecturas obtenidas de {len(dispositivos)} vehículos.")

    print("Buscando caídas sospechosas...")
    df_sospechosos = detectar_caidas_sospechosas(df_combustible)
    print(f"  {len(df_sospechosos)} eventos sospechosos encontrados.")

    print("Cargando zonas de Geotab...")
    zonas = obtener_zonas(client)
    print(f"  {len(zonas)} zonas cargadas.")

    print("Cruzando cada evento con la ubicación real del vehículo (esto puede tardar unos minutos)...")
    df_sospechosos = agregar_ubicacion_a_eventos(client, df_sospechosos, zonas)

    print("Trayendo odómetro para calcular cuánto se movió cada vehículo en cada evento...")
    df_odometro = extraer_odometro(client, dispositivos, fecha_inicio, fecha_fin)
    df_sospechosos = agregar_distancia_recorrida(df_sospechosos, df_odometro)

    print("Trayendo estado de ignición (encendido/apagado) para cada evento...")
    df_ignicion = extraer_ignicion(client, dispositivos, fecha_inicio, fecha_fin)
    df_sospechosos = agregar_estado_ignicion(df_sospechosos, df_ignicion)

    print("Intentando obtener altitud para aproximar la pendiente del terreno (opcional)...")
    df_altitud = extraer_altitud(client, dispositivos, fecha_inicio, fecha_fin)
    df_sospechosos = agregar_pendiente_aproximada(df_sospechosos, df_altitud)
    if 'Pendiente_Aprox_Pct' in df_sospechosos.columns and not df_sospechosos['Pendiente_Aprox_Pct'].notna().any():
        df_sospechosos = df_sospechosos.drop(columns=['Pendiente_Aprox_Pct'])

    print("Clasificando cada evento (robo probable / revisar sensor / posible ruido)...")
    df_sospechosos = clasificar_eventos(df_sospechosos)

    df_sospechosos = agregar_direccion_aproximada(df_sospechosos)

    resumen = resumen_por_vehiculo(df_sospechosos)
    print("\n=== Resumen por vehículo (de más a menos eventos) ===")
    print(resumen.to_string(index=False))

    if 'Categoria_Evento' in df_sospechosos.columns:
        print("\n=== Eventos por categoría ===")
        print(df_sospechosos['Categoria_Evento'].value_counts().to_string())

    if 'Cerca_Proveedor_Sospechoso' in df_sospechosos.columns:
        total_cerca = df_sospechosos['Cerca_Proveedor_Sospechoso'].sum()
        print(f"\n>>> {total_cerca} de {len(df_sospechosos)} eventos ocurrieron cerca de "
              f"alguna zona sospechosa ({', '.join(ZONAS_SOSPECHOSAS)}).")

    if 'Vehiculo_Practicamente_Detenido' in df_sospechosos.columns:
        total_detenido = df_sospechosos['Vehiculo_Practicamente_Detenido'].sum()
        print(f">>> {total_detenido} de {len(df_sospechosos)} eventos ocurrieron con el vehículo "
              f"prácticamente detenido (menos de {DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR} km recorridos) — "
              f"estos son los más creíbles como robo real, ya que el consumo normal de motor no explica la caída.")

    if 'Vehiculo_Apagado' in df_sospechosos.columns:
        total_apagado = df_sospechosos['Vehiculo_Apagado'].sum()
        print(f">>> {total_apagado} de {len(df_sospechosos)} eventos ocurrieron con el motor "
              f"APAGADO — la señal más fuerte de robo, ya que no puede haber consumo normal sin el motor encendido.")

    # Exporta el detalle completo a Excel, con formato para que sea claro para cualquiera
    if not df_sospechosos.empty:
        df_sospechosos = df_sospechosos.drop(columns=['DeviceId'])
        # Excel no admite datetimes con zona horaria, hay que quitarla antes de guardar
        for col in ['Fecha_Anterior', 'Fecha_Hora', 'Hora_Ultimo_Encendido', 'Hora_Ultimo_Apagado']:
            if col in df_sospechosos.columns:
                df_sospechosos[col] = pd.to_datetime(df_sospechosos[col]).dt.tz_localize(None)
        for col in ['Primer_Evento', 'Ultimo_Evento']:
            resumen[col] = resumen[col].dt.tz_localize(None)

        if 'Cerca_Proveedor_Sospechoso' in df_sospechosos.columns:
            df_sospechosos['Cerca_Proveedor_Sospechoso'] = df_sospechosos[
                'Cerca_Proveedor_Sospechoso'].map({True: 'Sí', False: 'No'})
        if 'Vehiculo_Practicamente_Detenido' in df_sospechosos.columns:
            df_sospechosos['Vehiculo_Practicamente_Detenido'] = df_sospechosos[
                'Vehiculo_Practicamente_Detenido'].map({True: 'Sí', False: 'No'})
        if 'Vehiculo_Apagado' in df_sospechosos.columns:
            df_sospechosos['Vehiculo_Apagado'] = df_sospechosos[
                'Vehiculo_Apagado'].map({True: 'Sí', False: 'No'})

        nombre_archivo = (f"posibles_robos_combustible_{fecha_inicio.date()}_{fecha_fin.date()}_"
                          f"{datetime.datetime.now().strftime('%H%M%S')}.xlsx")
        generar_excel_formateado(nombre_archivo, resumen, df_sospechosos)
        print(f"\nDetalle exportado a: {nombre_archivo}")