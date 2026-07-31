"""
Detección de posibles robos de combustible.

Busca caídas bruscas del nivel de combustible (%) en ventanas cortas de tiempo,
por vehículo, para identificar si el patrón se repite en el mismo móvil o
en varios distintos.

Requiere las mismas credenciales de Geotab que ya usa el tablero (variables
de entorno GEOTAB_USUARIO, GEOTAB_CONTRASENA, GEOTAB_DATABASE, GEOTAB_SERVER,
o un archivo .env en la misma carpeta).
"""

import os
import datetime
import mygeotab
import pandas as pd
from dotenv import load_dotenv

load_dotenv()  # carga las variables desde el archivo .env en la misma carpeta

NOMBRE_ZONA_PROVEEDOR_SOSPECHOSO = 'GUERRERO'  # coincide con "CASA TEJAR BASCULA GUERRERO"


def obtener_zonas(client):
    zonas_raw = client.get('Zone')
    zonas = []
    for z in zonas_raw:
        if isinstance(z, dict) and z.get('points'):
            poligono = [(p['x'], p['y']) for p in z['points']]
            zonas.append({'nombre': z.get('name', 'Zona sin nombre'), 'poligono': poligono})
    return zonas


def punto_en_poligono(lon, lat, poligono):
    n = len(poligono)
    dentro = False
    x1, y1 = poligono[0]
    for i in range(1, n + 1):
        x2, y2 = poligono[i % n]
        if lat > min(y1, y2) and lat <= max(y1, y2) and lon <= max(x1, x2):
            if y1 != y2:
                x_interseccion = (lat - y1) * (x2 - x1) / (y2 - y1) + x1
            if x1 == x2 or lon <= x_interseccion:
                dentro = not dentro
        x1, y1 = x2, y2
    return dentro


def determinar_zona(lon, lat, zonas):
    if lon is None or lat is None:
        return 'Sin ubicación'
    for z in zonas:
        if punto_en_poligono(lon, lat, z['poligono']):
            return z['nombre']
    return 'Fuera de zonas definidas'


def obtener_ubicacion_en_momento(client, device_id, momento):
    """Trae el punto GPS (LogRecord) más cercano a 'momento' para ese vehículo,
    buscando en una ventana corta (+/- 2 min) para no traer datos de más."""
    desde = (momento - datetime.timedelta(minutes=2)).isoformat()
    hasta = (momento + datetime.timedelta(minutes=2)).isoformat()
    registros = client.get('LogRecord', search={
        'deviceSearch': {'id': device_id},
        'fromDate': desde,
        'toDate': hasta,
    })
    if not registros:
        return None, None
    mas_cercano = min(registros, key=lambda r: abs(pd.to_datetime(r['dateTime']) - momento))
    return mas_cercano.get('longitude'), mas_cercano.get('latitude')


# --- Parámetros del análisis (ajusta a tu gusto) ---
UMBRAL_CAIDA_PCT = 15       # puntos porcentuales de caída para considerarla sospechosa
VENTANA_MAXIMA_MINUTOS = 30  # la caída debe ocurrir en máximo estos minutos
DIAS_A_ANALIZAR = 30         # cuántos días hacia atrás analizar

ID_DIAGNOSTICO_COMBUSTIBLE = 'DiagnosticFuelLevelId'
ID_DIAGNOSTICO_ODOMETRO = 'DiagnosticOdometerId'  # Geotab lo entrega en metros
DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR = 1.0  # por debajo de esto, se considera "prácticamente detenido"

REFERENCIA_MOTOR_POR_MARCA = {
    "volkswagen": "ISF 3.8",
    "volskwagen": "ISF 3.8",
    "mercedes": "OM926",
    "international": "L9",
    "foton": "X12",
    "kenworth": "ISM 11",
}

CIUDAD_A_FILTRAR = 'Bogotá'  # Cambia esto (o pon None) si quieres analizar otra ciudad o todas


def es_grupo_marca(nombre):
    nombre_l = nombre.strip().lower()
    return any(marca in nombre_l for marca in REFERENCIA_MOTOR_POR_MARCA)


def normalizar_ciudad(nombre):
    n = nombre.upper()
    if 'BOGOTA' in n or 'BOGOTÁ' in n:
        return 'Bogotá'
    if 'CALI' in n:
        return 'Cali'
    if 'VALLE' in n:
        return 'Valle'
    return nombre.strip()


def obtener_mapa_grupos(client):
    """Recorre la jerarquía de grupos de Geotab para saber a qué ciudad pertenece cada uno,
    exactamente igual a como lo hace el tablero de Streamlit."""
    todos_grupos = client.get('Group')
    grupos_por_id = {g.get('id'): g for g in todos_grupos if isinstance(g, dict)}
    raiz = next((g for g in todos_grupos if g.get('name', '').strip().startswith('*')), None)
    if not raiz:
        return {}

    mapa = {}

    def obtener_id(referencia):
        return referencia['id'] if isinstance(referencia, dict) else referencia

    def recorrer(grupo_id, ciudad_actual):
        grupo_completo = grupos_por_id.get(grupo_id)
        if not grupo_completo:
            return
        mapa[grupo_id] = {'nombre': grupo_completo.get('name', ''), 'ciudad': ciudad_actual}
        for hijo in (grupo_completo.get('children') or []):
            recorrer(obtener_id(hijo), ciudad_actual)

    for hijo_raiz in (raiz.get('children') or []):
        hijo_id = obtener_id(hijo_raiz)
        hijo_completo = grupos_por_id.get(hijo_id, {})
        nombre = hijo_completo.get('name', '').strip()
        if es_grupo_marca(nombre):
            recorrer(hijo_id, 'Sin ciudad asignada')
        else:
            ciudad = normalizar_ciudad(nombre)
            recorrer(hijo_id, ciudad)
    return mapa


def resolver_ciudad(vehiculo, mapa_grupos):
    for g in (vehiculo.get('groups') or []):
        gid = g['id'] if isinstance(g, dict) else g
        info = mapa_grupos.get(gid)
        if info and info['ciudad'] and info['ciudad'] != 'Sin ciudad asignada':
            return info['ciudad']
    return 'Sin ciudad asignada'


def conectar():
    usuario = os.environ["GEOTAB_USUARIO"]
    contrasena = os.environ["GEOTAB_CONTRASENA"]
    base_de_datos = os.environ["GEOTAB_DATABASE"]
    servidor = os.environ.get("GEOTAB_SERVER", "my.geotab.com")
    client = mygeotab.API(username=usuario, password=contrasena,
                           database=base_de_datos, server=servidor)
    client.authenticate()
    return client


def obtener_vehiculos(client):
    dispositivos = client.get('Device')
    mapa_grupos = obtener_mapa_grupos(client)

    resultado = {}
    for d in dispositivos:
        ciudad = resolver_ciudad(d, mapa_grupos)
        if CIUDAD_A_FILTRAR is not None and ciudad != CIUDAD_A_FILTRAR:
            continue
        resultado[d['id']] = d.get('name', d['id'])

    print(f"  {len(resultado)} de {len(dispositivos)} vehículos pertenecen a '{CIUDAD_A_FILTRAR}'.")
    return resultado


def extraer_niveles_combustible(client, dispositivos, f_inicio, f_fin):
    """Trae el nivel de combustible (%) de todos los vehículos en el rango dado."""
    calls = []
    for device_id in dispositivos:
        calls.append(('Get', {
            'typeName': 'StatusData',
            'search': {
                'deviceSearch': {'id': device_id},
                'diagnosticSearch': {'id': ID_DIAGNOSTICO_COMBUSTIBLE},
                'fromDate': f_inicio.isoformat(),
                'toDate': f_fin.isoformat(),
            }
        }))

    resultados = client.multi_call(calls)

    filas = []
    for device_id, resultado in zip(dispositivos.keys(), resultados):
        for registro in resultado:
            filas.append({
                'Vehiculo': dispositivos[device_id],
                'DeviceId': device_id,
                'Fecha_Hora': pd.to_datetime(registro['dateTime']),
                'Nivel_Combustible': registro['data'],
            })

    if not filas:
        return pd.DataFrame(columns=['Vehiculo', 'DeviceId', 'Fecha_Hora', 'Nivel_Combustible'])

    return pd.DataFrame(filas)


def extraer_odometro(client, dispositivos, f_inicio, f_fin):
    """Trae el odómetro (metros) de todos los vehículos en el rango dado, para poder
    calcular cuánto se movió cada uno entre dos lecturas de combustible."""
    calls = []
    for device_id in dispositivos:
        calls.append(('Get', {
            'typeName': 'StatusData',
            'search': {
                'deviceSearch': {'id': device_id},
                'diagnosticSearch': {'id': ID_DIAGNOSTICO_ODOMETRO},
                'fromDate': f_inicio.isoformat(),
                'toDate': f_fin.isoformat(),
            }
        }))

    resultados = client.multi_call(calls)

    filas = []
    for device_id, resultado in zip(dispositivos.keys(), resultados):
        for registro in resultado:
            filas.append({
                'Vehiculo': dispositivos[device_id],
                'Fecha_Hora': pd.to_datetime(registro['dateTime']),
                'Odometro_km': registro['data'] / 1000,
            })

    if not filas:
        return pd.DataFrame(columns=['Vehiculo', 'Fecha_Hora', 'Odometro_km'])

    return pd.DataFrame(filas).sort_values(['Vehiculo', 'Fecha_Hora'])


def agregar_distancia_recorrida(df_sospechosos, df_odometro):
    """Para cada evento sospechoso, calcula cuántos km recorrió el vehículo entre la
    lectura anterior y la lectura donde se detectó la caída, usando el odómetro más
    cercano disponible en cada momento (asof merge)."""
    if df_sospechosos.empty or df_odometro.empty:
        df_sospechosos = df_sospechosos.copy()
        df_sospechosos['Distancia_Recorrida_km'] = None
        df_sospechosos['Vehiculo_Practicamente_Detenido'] = None
        return df_sospechosos

    df = df_sospechosos.sort_values('Fecha_Anterior').copy()
    odo = df_odometro.sort_values('Fecha_Hora')

    antes = pd.merge_asof(
        df[['Vehiculo', 'Fecha_Anterior']].sort_values('Fecha_Anterior'),
        odo, left_on='Fecha_Anterior', right_on='Fecha_Hora',
        by='Vehiculo', direction='backward'
    )['Odometro_km']

    df_ord_despues = df.sort_values('Fecha_Hora')
    despues = pd.merge_asof(
        df_ord_despues[['Vehiculo', 'Fecha_Hora']],
        odo, on='Fecha_Hora', by='Vehiculo', direction='backward'
    )['Odometro_km']

    df = df.sort_values('Fecha_Anterior')
    df['Odometro_Antes_km'] = antes.values
    df = df.sort_values('Fecha_Hora')
    df['Odometro_Despues_km'] = despues.values

    df['Distancia_Recorrida_km'] = (df['Odometro_Despues_km'] - df['Odometro_Antes_km']).round(2)
    df['Vehiculo_Practicamente_Detenido'] = df['Distancia_Recorrida_km'] < DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR

    return df.drop(columns=['Odometro_Antes_km', 'Odometro_Despues_km'])


def detectar_caidas_sospechosas(df_combustible, umbral_pct=UMBRAL_CAIDA_PCT,
                                 ventana_min=VENTANA_MAXIMA_MINUTOS):
    """Marca como sospechosa cualquier caída >= umbral_pct en <= ventana_min minutos,
    comparando cada lectura contra la inmediatamente anterior del mismo vehículo."""
    if df_combustible.empty:
        return pd.DataFrame()

    df = df_combustible.sort_values(['Vehiculo', 'Fecha_Hora']).copy()
    df['Nivel_Anterior'] = df.groupby('Vehiculo')['Nivel_Combustible'].shift(1)
    df['Fecha_Anterior'] = df.groupby('Vehiculo')['Fecha_Hora'].shift(1)

    df['Caida_Pct'] = df['Nivel_Anterior'] - df['Nivel_Combustible']
    df['Minutos_Transcurridos'] = (df['Fecha_Hora'] - df['Fecha_Anterior']).dt.total_seconds() / 60

    sospechosos = df[
        (df['Caida_Pct'] >= umbral_pct) &
        (df['Minutos_Transcurridos'] <= ventana_min) &
        (df['Minutos_Transcurridos'] > 0)
    ].copy()

    return sospechosos[[
        'Vehiculo', 'DeviceId', 'Fecha_Anterior', 'Fecha_Hora', 'Nivel_Anterior',
        'Nivel_Combustible', 'Caida_Pct', 'Minutos_Transcurridos'
    ]].sort_values('Caida_Pct', ascending=False)


def agregar_ubicacion_a_eventos(client, df_sospechosos, zonas):
    """Para cada evento sospechoso, busca la posición GPS real del vehículo en ese
    momento y determina en qué zona de Geotab cae (ej. el taller/proveedor)."""
    if df_sospechosos.empty:
        return df_sospechosos

    zonas_detectadas = []
    cerca_proveedor = []
    total = len(df_sospechosos)
    for i, fila in enumerate(df_sospechosos.itertuples(), start=1):
        if i % 25 == 0 or i == total:
            print(f"  Validando ubicación: {i}/{total}...")
        lon, lat = obtener_ubicacion_en_momento(client, fila.DeviceId, fila.Fecha_Hora)
        zona = determinar_zona(lon, lat, zonas)
        zonas_detectadas.append(zona)
        cerca_proveedor.append(NOMBRE_ZONA_PROVEEDOR_SOSPECHOSO.upper() in zona.upper())

    df_sospechosos = df_sospechosos.copy()
    df_sospechosos['Zona_Detectada'] = zonas_detectadas
    df_sospechosos['Cerca_Proveedor_Sospechoso'] = cerca_proveedor
    return df_sospechosos


def resumen_por_vehiculo(df_sospechosos):
    """Cuenta cuántos eventos sospechosos tiene cada vehículo, para ver si se repite."""
    if df_sospechosos.empty:
        return pd.DataFrame()

    agregaciones = {
        'Eventos_Sospechosos': ('Caida_Pct', 'count'),
        'Caida_Promedio_Pct': ('Caida_Pct', 'mean'),
        'Caida_Maxima_Pct': ('Caida_Pct', 'max'),
        'Primer_Evento': ('Fecha_Hora', 'min'),
        'Ultimo_Evento': ('Fecha_Hora', 'max'),
    }
    if 'Cerca_Proveedor_Sospechoso' in df_sospechosos.columns:
        agregaciones['Eventos_Cerca_Proveedor'] = ('Cerca_Proveedor_Sospechoso', 'sum')
    if 'Vehiculo_Practicamente_Detenido' in df_sospechosos.columns:
        agregaciones['Eventos_Vehiculo_Detenido'] = ('Vehiculo_Practicamente_Detenido', 'sum')

    resumen = df_sospechosos.groupby('Vehiculo').agg(**agregaciones).sort_values(
        'Eventos_Sospechosos', ascending=False
    )

    return resumen.reset_index()


def generar_excel_formateado(nombre_archivo, resumen, detalle):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLOR_ENCABEZADO = "1F4E4A"   # verde oscuro, igual al del tablero
    COLOR_TEXTO_ENCABEZADO = "FFFFFF"
    COLOR_RESALTADO_PROVEEDOR = "FFF3B0"   # amarillo suave: cerca del proveedor sospechoso
    COLOR_RESALTADO_DETENIDO = "F4B4AE"    # rojo suave: caída sin movimiento del vehículo (más grave)
    COLOR_BORDE = "D9D9D9"

    fuente_encabezado = Font(bold=True, color=COLOR_TEXTO_ENCABEZADO, size=11)
    relleno_encabezado = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
    relleno_proveedor = PatternFill("solid", fgColor=COLOR_RESALTADO_PROVEEDOR)
    relleno_detenido = PatternFill("solid", fgColor=COLOR_RESALTADO_DETENIDO)
    borde_delgado = Border(*(Side(style='thin', color=COLOR_BORDE) for _ in range(4)))
    alineacion_centrada = Alignment(horizontal='center', vertical='center')

    wb = Workbook()

    # --- Hoja 1: Léame, explicando el análisis en lenguaje sencillo ---
    ws_intro = wb.active
    ws_intro.title = "Léame"
    ws_intro.column_dimensions['A'].width = 100
    lineas_intro = [
        ("ANÁLISIS DE POSIBLE ROBO DE COMBUSTIBLE", True, 14),
        ("", False, 11),
        (f"Periodo analizado: {fecha_inicio.date()} al {fecha_fin.date()}", False, 11),
        (f"Ciudad: {CIUDAD_A_FILTRAR or 'Todas'}", False, 11),
        ("", False, 11),
        ("¿Qué hace este análisis?", True, 12),
        ("Revisa el nivel de combustible (%) de cada vehículo y marca como 'sospechoso'", False, 11),
        (f"cualquier caída de {UMBRAL_CAIDA_PCT} puntos porcentuales o más, ocurrida en", False, 11),
        (f"{VENTANA_MAXIMA_MINUTOS} minutos o menos.", False, 11),
        ("", False, 11),
        ("¿Cómo leer la hoja 'Resumen por vehiculo'?", True, 12),
        ("Cada fila es un vehículo. Entre más eventos sospechosos tenga, más veces se le", False, 11),
        ("detectó una caída brusca de combustible en el periodo analizado.", False, 11),
        ("", False, 11),
        ("¿Cómo leer la hoja 'Detalle de eventos'?", True, 12),
        ("Cada fila es una caída puntual detectada, con fecha, hora, cuánto bajó el", False, 11),
        ("combustible y en cuántos minutos. La columna 'Zona_Detectada' indica en qué", False, 11),
        ("zona se encontraba el vehículo, y 'Distancia_Recorrida_km' cuánto se movió", False, 11),
        ("entre esas dos lecturas.", False, 11),
        ("", False, 11),
        ("Colores de las filas:", True, 12),
        ("🔴 Rojo: el vehículo casi no se movió (menos de "
         f"{DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR} km) mientras bajaba el", False, 11),
        ("combustible — el consumo normal de motor NO explica esta caída. Son los casos", False, 11),
        ("más creíbles como robo real y los que ameritan revisar primero.", False, 11),
        ("🟡 Amarillo: ocurrió cerca de la zona del proveedor bajo sospecha.", False, 11),
        ("", False, 11),
        ("⚠️ Importante:", True, 12),
        ("Una caída detectada NO confirma un robo por sí sola — puede deberse también a", False, 11),
        ("ruido del sensor de combustible. Prioriza revisar las filas en rojo primero,", False, 11),
        ("y valida contra el historial de viajes real antes de tomar acción.", False, 11),
    ]
    for i, (texto, negrita, tamano) in enumerate(lineas_intro, start=1):
        celda = ws_intro.cell(row=i, column=1, value=texto)
        celda.font = Font(bold=negrita, size=tamano,
                           color=COLOR_ENCABEZADO if negrita else "000000")
        celda.alignment = Alignment(wrap_text=True, vertical='center')

    # --- Hojas de datos ---
    for nombre_hoja, df in [("Resumen por vehiculo", resumen), ("Detalle de eventos", detalle)]:
        ws = wb.create_sheet(nombre_hoja)
        ws.append(list(df.columns))
        for celda in ws[1]:
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado
            celda.alignment = alineacion_centrada
            celda.border = borde_delgado
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        col_proveedor = None
        if 'Cerca_Proveedor_Sospechoso' in df.columns:
            col_proveedor = list(df.columns).index('Cerca_Proveedor_Sospechoso') + 1
        col_detenido = None
        if 'Vehiculo_Practicamente_Detenido' in df.columns:
            col_detenido = list(df.columns).index('Vehiculo_Practicamente_Detenido') + 1

        for fila in df.itertuples(index=False):
            ws.append(list(fila))
            fila_actual = ws.max_row
            es_detenido = col_detenido and ws.cell(row=fila_actual, column=col_detenido).value == 'Sí'
            es_proveedor = col_proveedor and ws.cell(row=fila_actual, column=col_proveedor).value == 'Sí'
            for celda in ws[fila_actual]:
                celda.border = borde_delgado
                if es_detenido:
                    celda.fill = relleno_detenido
                elif es_proveedor:
                    celda.fill = relleno_proveedor

        # Ajustar ancho de columnas según el contenido
        for i, col in enumerate(df.columns, start=1):
            letra = get_column_letter(i)
            ancho_maximo = max(
                [len(str(col))] + [len(str(v)) for v in df[col].astype(str).values]
            )
            ws.column_dimensions[letra].width = min(ancho_maximo + 4, 45)

    wb.save(nombre_archivo)


if __name__ == '__main__':
    print("Conectando a Geotab...")
    client = conectar()

    fecha_fin = datetime.datetime.now(datetime.UTC)
    fecha_inicio = fecha_fin - datetime.timedelta(days=DIAS_A_ANALIZAR)

    print(f"Trayendo niveles de combustible de los últimos {DIAS_A_ANALIZAR} días...")
    dispositivos = obtener_vehiculos(client)
    df_combustible = extraer_niveles_combustible(client, dispositivos, fecha_inicio, fecha_fin)
    print(f"  {len(df_combustible)} lecturas obtenidas de {len(dispositivos)} vehículos.")

    print("Buscando caídas sospechosas...")
    df_sospechosos = detectar_caidas_sospechosas(df_combustible)
    print(f"  {len(df_sospechosos)} eventos sospechosos encontrados.")

    print("Cargando zonas de Geotab...")
    zonas = obtener_zonas(client)
    print(f"  {len(zonas)} zonas cargadas.")

    print("Cruzando cada evento con la ubicación real del vehículo (esto puede tardar unos minutos)...")
    df_sospechosos = agregar_ubicacion_a_eventos(client, df_sospechosos, zonas)

    print("Trayendo odómetro para calcular cuánto se movió cada vehículo en cada evento...")
    df_odometro = extraer_odometro(client, dispositivos, fecha_inicio, fecha_fin)
    df_sospechosos = agregar_distancia_recorrida(df_sospechosos, df_odometro)

    resumen = resumen_por_vehiculo(df_sospechosos)
    print("\n=== Resumen por vehículo (de más a menos eventos) ===")
    print(resumen.to_string(index=False))

    if 'Cerca_Proveedor_Sospechoso' in df_sospechosos.columns:
        total_cerca = df_sospechosos['Cerca_Proveedor_Sospechoso'].sum()
        print(f"\n>>> {total_cerca} de {len(df_sospechosos)} eventos ocurrieron cerca de "
              f"la zona del proveedor sospechoso ('{NOMBRE_ZONA_PROVEEDOR_SOSPECHOSO}').")

    if 'Vehiculo_Practicamente_Detenido' in df_sospechosos.columns:
        total_detenido = df_sospechosos['Vehiculo_Practicamente_Detenido'].sum()
        print(f">>> {total_detenido} de {len(df_sospechosos)} eventos ocurrieron con el vehículo "
              f"prácticamente detenido (menos de {DISTANCIA_MINIMA_KM_PARA_JUSTIFICAR} km recorridos) — "
              f"estos son los más creíbles como robo real, ya que el consumo normal de motor no explica la caída.")

    # Exporta el detalle completo a Excel, con formato para que sea claro para cualquiera
    if not df_sospechosos.empty:
        df_sospechosos = df_sospechosos.drop(columns=['DeviceId'])
        # Excel no admite datetimes con zona horaria, hay que quitarla antes de guardar
        for col in ['Fecha_Anterior', 'Fecha_Hora']:
            df_sospechosos[col] = df_sospechosos[col].dt.tz_localize(None)
        for col in ['Primer_Evento', 'Ultimo_Evento']:
            resumen[col] = resumen[col].dt.tz_localize(None)

        if 'Cerca_Proveedor_Sospechoso' in df_sospechosos.columns:
            df_sospechosos['Cerca_Proveedor_Sospechoso'] = df_sospechosos[
                'Cerca_Proveedor_Sospechoso'].map({True: 'Sí', False: 'No'})
        if 'Vehiculo_Practicamente_Detenido' in df_sospechosos.columns:
            df_sospechosos['Vehiculo_Practicamente_Detenido'] = df_sospechosos[
                'Vehiculo_Practicamente_Detenido'].map({True: 'Sí', False: 'No'})

        nombre_archivo = (f"posibles_robos_combustible_{fecha_inicio.date()}_{fecha_fin.date()}_"
                          f"{datetime.datetime.now().strftime('%H%M%S')}.xlsx")
        generar_excel_formateado(nombre_archivo, resumen, df_sospechosos)
        print(f"\nDetalle exportado a: {nombre_archivo}")